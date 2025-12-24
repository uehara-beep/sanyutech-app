from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func
from pydantic import BaseModel
from typing import Optional, List
from datetime import date, datetime
import httpx
from database import engine, get_db, Base
from models import (
    Project, Cost, Billing, FixedCost, Client, Vendor, Material,
    Machine, WorkType, Settings, BudgetDetail,
    ExpenseCategory, FuelPrice, Expense, ExpenseReceipt,
    ProjectWorkType, WorkTypeDetail, MonthlyProgress, Receivable, Payable,
    Estimate, Budget, Approval, Notification, Worker, Assignment,
    KYReport, KYSignature, InventoryItem, InventoryTransaction,
    Vehicle, VehicleLog, Schedule, Attendance, Subcontractor,
    Order, Checklist, EmergencyContact, Equipment, Template,
    Drawing, DrawingPin, SitePhoto, BlackboardTemplate,
    Inspection, InspectionItem, Correction,
    WorkerRegistration, SafetyTraining, Qualification,
    Message, MessageRead, User, Permission, IntegrationSetting,
    DailyReport, DocumentSend, CompanySettings,
    LineWorksSettings, LineWorksUser, LineWorksNotification, LineWorksLog,
    BusinessCard, QuoteDocument, QuoteItem,
    Member, HotelRequest
)
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os

Base.metadata.create_all(bind=engine)

app = FastAPI(title="S-BASE API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Pydantic Models
class ProjectCreate(BaseModel):
    code: str
    name: str
    client: Optional[str] = None
    status: Optional[str] = "見積中"
    order_type: Optional[str] = "一次請"
    prefecture: Optional[str] = None
    probability: Optional[str] = "確定"
    order_amount: Optional[int] = 0
    budget_amount: Optional[int] = 0
    tax_rate: Optional[float] = 0.1
    period: Optional[str] = None
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    sales_person: Optional[str] = None
    site_person: Optional[str] = None
    address: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None

class CostCreate(BaseModel):
    project_id: int
    date: date
    category: str
    work_type: Optional[str] = None
    vendor: Optional[str] = None
    description: Optional[str] = None
    quantity: Optional[float] = 1
    unit: Optional[str] = None
    unit_price: Optional[int] = 0
    amount: Optional[int] = 0

class ClientCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    address: Optional[str] = None
    closing_day: Optional[int] = 25
    payment_day: Optional[int] = 25
    payment_month_offset: Optional[int] = 1

class VendorCreate(BaseModel):
    name: str
    category: Optional[str] = None
    default_price: Optional[int] = 0
    unit: Optional[str] = None
    phone: Optional[str] = None
    closing_day: Optional[int] = 25
    payment_day: Optional[int] = 25
    payment_month_offset: Optional[int] = 1

class MaterialCreate(BaseModel):
    name: str
    category: Optional[str] = None
    unit: Optional[str] = None
    unit_price: Optional[int] = 0
    vendor: Optional[str] = None

class MachineCreate(BaseModel):
    name: str
    category: Optional[str] = None
    unit: Optional[str] = "台"
    unit_price: Optional[int] = 0
    vendor: Optional[str] = None

class WorkTypeCreate(BaseModel):
    name: str
    category: Optional[str] = None

class SettingUpdate(BaseModel):
    key: str
    value: str

class MonthlyProgressCreate(BaseModel):
    project_id: int
    year_month: str
    progress_amount: Optional[int] = 0
    progress_rate: Optional[float] = 0
    cost_amount: Optional[int] = 0
    gross_profit: Optional[int] = 0
    gross_profit_rate: Optional[float] = 0
    note: Optional[str] = None

class ReceivableCreate(BaseModel):
    project_id: int
    progress_id: Optional[int] = None
    client_name: str
    description: Optional[str] = None
    amount: Optional[int] = 0
    billing_date: Optional[date] = None
    expected_date: date
    actual_date: Optional[date] = None
    status: Optional[str] = "予定"
    note: Optional[str] = None

class PayableCreate(BaseModel):
    project_id: Optional[int] = None
    cost_id: Optional[int] = None
    vendor_name: str
    category: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[int] = 0
    invoice_date: Optional[date] = None
    expected_date: date
    actual_date: Optional[date] = None
    status: Optional[str] = "予定"
    note: Optional[str] = None

class EstimateCreate(BaseModel):
    project_id: int
    work_type: Optional[str] = None
    description: Optional[str] = None
    quantity: Optional[float] = 0
    unit: Optional[str] = None
    unit_price: Optional[float] = 0
    amount: Optional[float] = 0
    rate: Optional[float] = 1.0

class BudgetCreate(BaseModel):
    project_id: int
    work_type: Optional[str] = None
    category: Optional[str] = None
    vendor: Optional[str] = None
    description: Optional[str] = None
    quantity: Optional[float] = 0
    unit: Optional[str] = None
    unit_price: Optional[float] = 0
    amount: Optional[float] = 0

class ApprovalUpdate(BaseModel):
    status: str
    approved_by: Optional[str] = None
    comment: Optional[str] = None

class WorkerCreate(BaseModel):
    name: str
    team: Optional[str] = None
    employment_type: Optional[str] = None
    daily_rate: Optional[float] = 0
    phone: Optional[str] = None
    is_active: Optional[bool] = True

class AssignmentCreate(BaseModel):
    date: date
    project_id: Optional[int] = None
    worker_id: Optional[int] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    note: Optional[str] = None

class KYReportCreate(BaseModel):
    project_id: Optional[int] = None
    date: date
    work_content: Optional[str] = None
    hazards: Optional[str] = None
    countermeasures: Optional[str] = None
    participants: Optional[str] = None
    photo_path: Optional[str] = None
    created_by: Optional[str] = None

class InventoryItemCreate(BaseModel):
    name: str
    category: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = 0
    min_quantity: Optional[float] = 0
    location: Optional[str] = None

class InventoryTransactionCreate(BaseModel):
    item_id: int
    type: str
    quantity: float
    project_id: Optional[int] = None
    date: Optional[date] = None
    note: Optional[str] = None

class VehicleCreate(BaseModel):
    name: str
    plate_number: Optional[str] = None
    type: Optional[str] = None
    inspection_date: Optional[date] = None
    insurance_date: Optional[date] = None

class VehicleLogCreate(BaseModel):
    vehicle_id: int
    date: Optional[date] = None
    driver: Optional[str] = None
    mileage: Optional[float] = None
    fuel_amount: Optional[float] = None
    fuel_cost: Optional[float] = None
    note: Optional[str] = None

class ScheduleCreate(BaseModel):
    project_id: int
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    progress_rate: Optional[float] = 0
    color: Optional[str] = "#3b82f6"

class AttendanceCreate(BaseModel):
    worker_id: int
    date: date
    check_in: Optional[datetime] = None
    check_out: Optional[datetime] = None
    project_id: Optional[int] = None
    overtime_hours: Optional[float] = 0
    note: Optional[str] = None

class SubcontractorCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    category: Optional[str] = None
    note: Optional[str] = None

class OrderCreate(BaseModel):
    project_id: Optional[int] = None
    vendor_id: Optional[int] = None
    item_name: Optional[str] = None
    quantity: Optional[float] = None
    unit: Optional[str] = None
    unit_price: Optional[float] = None
    total: Optional[float] = None
    status: Optional[str] = "draft"
    order_date: Optional[date] = None
    delivery_date: Optional[date] = None
    note: Optional[str] = None

class ChecklistCreate(BaseModel):
    project_id: Optional[int] = None
    name: Optional[str] = None
    items: Optional[str] = None
    completed_items: Optional[str] = None

class EmergencyContactCreate(BaseModel):
    name: str
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    priority: Optional[int] = 0

class EquipmentCreate(BaseModel):
    name: str
    category: Optional[str] = None
    status: Optional[str] = "available"
    current_project_id: Optional[int] = None
    note: Optional[str] = None

class TemplateCreate(BaseModel):
    name: str
    type: Optional[str] = None
    content: Optional[str] = None

class ExpenseCreate(BaseModel):
    project_id: int
    category_id: int
    expense_date: date
    amount: Optional[float] = None
    fuel_type: Optional[str] = None
    fuel_liter: Optional[float] = None
    store_name: Optional[str] = None
    memo: Optional[str] = None

class BillingCreate(BaseModel):
    project_id: int
    bill_type: str
    date: date
    amount: Optional[int] = 0
    done: Optional[int] = 0
    note: Optional[str] = None

class NotificationCreate(BaseModel):
    user_id: Optional[str] = None
    type: Optional[str] = None
    title: Optional[str] = None
    message: Optional[str] = None
    link: Optional[str] = None

# ========== Projects ==========
@app.get("/api/projects")
def get_projects(db: Session = Depends(get_db)):
    projects = db.query(Project).order_by(Project.id.desc()).all()
    result = []
    for p in projects:
        total_cost = db.query(func.sum(Cost.amount)).filter(Cost.project_id == p.id).scalar() or 0
        result.append({
            "id": p.id, "code": p.code, "name": p.name, "client": p.client, "status": p.status,
            "order_type": p.order_type, "prefecture": p.prefecture, "probability": p.probability,
            "order_amount": p.order_amount, "budget_amount": p.budget_amount, "tax_rate": p.tax_rate,
            "period": p.period, "sales_person": p.sales_person, "site_person": p.site_person,
            "address": p.address, "latitude": p.latitude, "longitude": p.longitude,
            "start_date": str(p.start_date) if p.start_date else None,
            "end_date": str(p.end_date) if p.end_date else None,
            "total_cost": total_cost,
            "contract_amount": p.order_amount,  # 互換性のため
            "actual_cost": total_cost,  # 互換性のため
            "sales_profit": (p.order_amount or 0) - (p.budget_amount or 0),
            "construction_profit": (p.budget_amount or 0) - total_cost,
        })
    return result

@app.get("/api/projects/{project_id}")
def get_project(project_id: int, db: Session = Depends(get_db)):
    """単一工事の取得"""
    p = db.query(Project).filter(Project.id == project_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Project not found")
    total_cost = db.query(func.sum(Cost.amount)).filter(Cost.project_id == p.id).scalar() or 0
    return {
        "id": p.id, "code": p.code, "name": p.name, "client": p.client, "status": p.status,
        "order_type": p.order_type, "prefecture": p.prefecture, "probability": p.probability,
        "order_amount": p.order_amount, "budget_amount": p.budget_amount, "tax_rate": p.tax_rate,
        "period": p.period, "start_date": str(p.start_date) if p.start_date else None,
        "end_date": str(p.end_date) if p.end_date else None,
        "sales_person": p.sales_person, "site_person": p.site_person,
        "address": p.address, "latitude": p.latitude, "longitude": p.longitude,
        "total_cost": total_cost,
        "contract_amount": p.order_amount,  # 互換性のため
        "sales_profit": (p.order_amount or 0) - (p.budget_amount or 0),
        "construction_profit": (p.budget_amount or 0) - total_cost,
        "created_at": str(p.created_at) if p.created_at else None,
        "updated_at": str(p.updated_at) if p.updated_at else None,
    }

@app.post("/api/projects")
def create_project(project: ProjectCreate, db: Session = Depends(get_db)):
    db_project = Project(**project.dict())
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    return db_project

@app.put("/api/projects/{project_id}")
def update_project(project_id: int, project: ProjectCreate, db: Session = Depends(get_db)):
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if not db_project:
        raise HTTPException(status_code=404, detail="Project not found")
    for key, value in project.dict().items():
        setattr(db_project, key, value)
    db.commit()
    return db_project

@app.delete("/api/projects/{project_id}")
def delete_project(project_id: int, db: Session = Depends(get_db)):
    db_project = db.query(Project).filter(Project.id == project_id).first()
    if db_project:
        db.query(Cost).filter(Cost.project_id == project_id).delete()
        db.delete(db_project)
        db.commit()
    return {"ok": True}

# ========== Costs ==========
@app.get("/api/costs")
def get_costs(project_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Cost)
    if project_id:
        query = query.filter(Cost.project_id == project_id)
    return query.order_by(Cost.date.desc()).all()

@app.post("/api/costs")
def create_cost(cost: CostCreate, db: Session = Depends(get_db)):
    db_cost = Cost(**cost.dict())
    db.add(db_cost)
    db.commit()
    db.refresh(db_cost)
    return db_cost

@app.delete("/api/costs/{cost_id}")
def delete_cost(cost_id: int, db: Session = Depends(get_db)):
    db_cost = db.query(Cost).filter(Cost.id == cost_id).first()
    if db_cost:
        db.delete(db_cost)
        db.commit()
    return {"ok": True}

# ========== Monthly Data ==========
@app.get("/api/projects/{project_id}/monthly")
def get_monthly_data(project_id: int, db: Session = Depends(get_db)):
    costs = db.query(Cost).filter(Cost.project_id == project_id).all()
    cost_by_month = {}
    for c in costs:
        if c.date:
            key = c.date.strftime("%Y-%m")
            if key not in cost_by_month:
                cost_by_month[key] = {"労務費": 0, "材料費": 0, "外注費": 0, "経費": 0, "合計": 0}
            cost_by_month[key][c.category] = cost_by_month[key].get(c.category, 0) + (c.amount or 0)
            cost_by_month[key]["合計"] += (c.amount or 0)
    return [{"month": k, **v} for k, v in sorted(cost_by_month.items())]

# ========== Work Type Stats ==========
@app.get("/api/projects/{project_id}/worktype-stats")
def get_worktype_stats(project_id: int, db: Session = Depends(get_db)):
    costs = db.query(Cost).filter(Cost.project_id == project_id).all()
    stats = {}
    for c in costs:
        wt = c.work_type or "未分類"
        if wt not in stats:
            stats[wt] = {"労務費": 0, "材料費": 0, "外注費": 0, "経費": 0, "合計": 0}
        stats[wt][c.category] = stats[wt].get(c.category, 0) + (c.amount or 0)
        stats[wt]["合計"] += (c.amount or 0)
    return [{"work_type": k, **v} for k, v in stats.items()]

# ========== Dashboard ==========
@app.get("/api/dashboard")
def get_dashboard(db: Session = Depends(get_db)):
    projects = db.query(Project).all()
    total_order = sum(p.order_amount or 0 for p in projects)
    total_budget = sum(p.budget_amount or 0 for p in projects)
    total_cost = db.query(func.sum(Cost.amount)).scalar() or 0
    
    confirmed = [p for p in projects if p.probability == "確定"]
    likely = [p for p in projects if p.probability == "見込み有"]
    
    return {
        "total_order": total_order,
        "total_budget": total_budget,
        "total_cost": total_cost,
        "total_profit": total_order - total_cost,
        "sales_profit": total_order - total_budget,
        "construction_profit": total_budget - total_cost,
        "project_count": len(projects),
        "confirmed_count": len(confirmed),
        "confirmed_amount": sum(p.order_amount or 0 for p in confirmed),
        "likely_count": len(likely),
        "likely_amount": sum(p.order_amount or 0 for p in likely),
    }

# ========== Clients (元請けマスタ) ==========
@app.get("/api/clients")
def get_clients(db: Session = Depends(get_db)):
    return db.query(Client).order_by(Client.name).all()

@app.post("/api/clients")
def create_client(client: ClientCreate, db: Session = Depends(get_db)):
    db_client = Client(**client.dict())
    db.add(db_client)
    db.commit()
    db.refresh(db_client)
    return db_client


@app.put("/api/clients/{client_id}")
def update_client(client_id: int, client: dict, db: Session = Depends(get_db)):
    db_client = db.query(Client).filter(Client.id == client_id).first()
    if db_client:
        db_client.name = client.get("name", db_client.name)
        db_client.contact_person = client.get("contact_person", db_client.contact_person)
        db_client.phone = client.get("phone", db_client.phone)
        db.commit()
        db.refresh(db_client)
    return db_client

@app.delete("/api/clients/{client_id}")
def delete_client(client_id: int, db: Session = Depends(get_db)):
    db_client = db.query(Client).filter(Client.id == client_id).first()
    if db_client:
        db.delete(db_client)
        db.commit()
    return {"ok": True}

# ========== Vendors (業者マスタ) ==========
@app.get("/api/vendors")
def get_vendors(db: Session = Depends(get_db)):
    return db.query(Vendor).order_by(Vendor.name).all()

@app.post("/api/vendors")
def create_vendor(vendor: VendorCreate, db: Session = Depends(get_db)):
    db_vendor = Vendor(**vendor.dict())
    db.add(db_vendor)
    db.commit()
    db.refresh(db_vendor)
    return db_vendor

@app.put("/api/vendors/{vendor_id}")
def update_vendor(vendor_id: int, vendor: VendorCreate, db: Session = Depends(get_db)):
    db_vendor = db.query(Vendor).filter(Vendor.id == vendor_id).first()
    if db_vendor:
        for key, value in vendor.dict().items():
            setattr(db_vendor, key, value)
        db.commit()
    return db_vendor

@app.delete("/api/vendors/{vendor_id}")
def delete_vendor(vendor_id: int, db: Session = Depends(get_db)):
    db_vendor = db.query(Vendor).filter(Vendor.id == vendor_id).first()
    if db_vendor:
        db.delete(db_vendor)
        db.commit()
    return {"ok": True}

# ========== Materials (材料マスタ) ==========
@app.get("/api/materials")
def get_materials(db: Session = Depends(get_db)):
    return db.query(Material).order_by(Material.name).all()

@app.post("/api/materials")
def create_material(material: MaterialCreate, db: Session = Depends(get_db)):
    db_material = Material(**material.dict())
    db.add(db_material)
    db.commit()
    db.refresh(db_material)
    return db_material

@app.put("/api/materials/{material_id}")
def update_material(material_id: int, material: MaterialCreate, db: Session = Depends(get_db)):
    db_material = db.query(Material).filter(Material.id == material_id).first()
    if db_material:
        for key, value in material.dict().items():
            setattr(db_material, key, value)
        db.commit()
    return db_material

@app.delete("/api/materials/{material_id}")
def delete_material(material_id: int, db: Session = Depends(get_db)):
    db_material = db.query(Material).filter(Material.id == material_id).first()
    if db_material:
        db.delete(db_material)
        db.commit()
    return {"ok": True}

# ========== Machines (機械マスタ) ==========
@app.get("/api/machines")
def get_machines(db: Session = Depends(get_db)):
    return db.query(Machine).order_by(Machine.name).all()

@app.post("/api/machines")
def create_machine(machine: MachineCreate, db: Session = Depends(get_db)):
    db_machine = Machine(**machine.dict())
    db.add(db_machine)
    db.commit()
    db.refresh(db_machine)
    return db_machine

@app.put("/api/machines/{machine_id}")
def update_machine(machine_id: int, machine: MachineCreate, db: Session = Depends(get_db)):
    db_machine = db.query(Machine).filter(Machine.id == machine_id).first()
    if db_machine:
        for key, value in machine.dict().items():
            setattr(db_machine, key, value)
        db.commit()
    return db_machine

@app.delete("/api/machines/{machine_id}")
def delete_machine(machine_id: int, db: Session = Depends(get_db)):
    db_machine = db.query(Machine).filter(Machine.id == machine_id).first()
    if db_machine:
        db.delete(db_machine)
        db.commit()
    return {"ok": True}

# ========== Work Types (工種マスタ) ==========
@app.get("/api/work-types")
def get_work_types(db: Session = Depends(get_db)):
    return db.query(WorkType).order_by(WorkType.name).all()

@app.post("/api/work-types")
def create_work_type(work_type: WorkTypeCreate, db: Session = Depends(get_db)):
    db_wt = WorkType(**work_type.dict())
    db.add(db_wt)
    db.commit()
    db.refresh(db_wt)
    return db_wt

@app.delete("/api/work-types/{wt_id}")
def delete_work_type(wt_id: int, db: Session = Depends(get_db)):
    db_wt = db.query(WorkType).filter(WorkType.id == wt_id).first()
    if db_wt:
        db.delete(db_wt)
        db.commit()
    return {"ok": True}

# ========== Settings ==========
@app.get("/api/settings")
def get_settings(db: Session = Depends(get_db)):
    settings = db.query(Settings).all()
    return {s.key: s.value for s in settings}

@app.post("/api/settings")
def update_setting(setting: SettingUpdate, db: Session = Depends(get_db)):
    db_setting = db.query(Settings).filter(Settings.key == setting.key).first()
    if db_setting:
        db_setting.value = setting.value
    else:
        db_setting = Settings(key=setting.key, value=setting.value)
        db.add(db_setting)
    db.commit()
    return {"ok": True}


# 予算明細 API
@app.get("/api/budget-details/{project_id}")
def get_budget_details(project_id: int, db: Session = Depends(get_db)):
    return db.query(BudgetDetail).filter(BudgetDetail.project_id == project_id).all()

@app.post("/api/budget-details")
def create_budget_detail(detail: dict, db: Session = Depends(get_db)):
    db_detail = BudgetDetail(
        project_id=detail.get("project_id"),
        category=detail.get("category"),
        work_type=detail.get("work_type"),
        vendor=detail.get("vendor"),
        description=detail.get("description"),
        quantity=detail.get("quantity", 0),
        unit=detail.get("unit"),
        unit_price=detail.get("unit_price", 0),
        amount=detail.get("quantity", 0) * detail.get("unit_price", 0)
    )
    db.add(db_detail)
    db.commit()
    db.refresh(db_detail)
    return db_detail

@app.delete("/api/budget-details/{detail_id}")
def delete_budget_detail(detail_id: int, db: Session = Depends(get_db)):
    db_detail = db.query(BudgetDetail).filter(BudgetDetail.id == detail_id).first()
    if db_detail:
        db.delete(db_detail)
        db.commit()
    return {"ok": True}

@app.put("/api/budget-details/{detail_id}")
def update_budget_detail(detail_id: int, detail: dict, db: Session = Depends(get_db)):
    db_detail = db.query(BudgetDetail).filter(BudgetDetail.id == detail_id).first()
    if db_detail:
        db_detail.category = detail.get("category", db_detail.category)
        db_detail.work_type = detail.get("work_type", db_detail.work_type)
        db_detail.vendor = detail.get("vendor", db_detail.vendor)
        db_detail.description = detail.get("description", db_detail.description)
        db_detail.quantity = detail.get("quantity", db_detail.quantity)
        db_detail.unit = detail.get("unit", db_detail.unit)
        db_detail.unit_price = detail.get("unit_price", db_detail.unit_price)
        db_detail.amount = detail.get("quantity", 0) * detail.get("unit_price", 0)
        db.commit()
        db.refresh(db_detail)
    return db_detail




# Excelファイルアップロード用
from fastapi import File, UploadFile
import openpyxl
from io import BytesIO

@app.post("/api/budget-details/upload/{project_id}")
async def upload_budget_excel(project_id: int, file: UploadFile = File(...), vendor: str = "", category: str = "外注費", db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    
    items = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 空行スキップ
            if not row or all(cell is None for cell in row):
                continue
            
            # 名称を探す（最初の文字列セル）
            name = None
            spec = ""
            qty = 0
            unit = "式"
            price = 0
            amount = 0
            note = ""
            
            # 列を解析
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                cell_str = str(cell).strip()
                
                # スキップする行
                skip_words = ['直接工事費', '諸経費', '機械回送費', '小計', '合計', '端数調整', '法定福利費', '労働災害', '内訳', '名称', '規格', '数量', '単位', '単価', '金額', '備考']
                if any(skip in cell_str for skip in skip_words):
                    name = None
                    break
                
                # 数値判定
                if isinstance(cell, (int, float)):
                    if name and qty == 0 and cell > 0:
                        qty = float(cell)
                    elif name and qty > 0 and price == 0 and cell > 100:
                        price = int(cell)
                    elif name and price > 0 and amount == 0:
                        amount = int(cell)
                elif isinstance(cell, str) and cell.strip():
                    if name is None:
                        name = cell.strip()
                    elif not spec:
                        spec = cell.strip()
                    elif cell in ['m2', 'ｍ2', 'm3', 'ｍ3', '式', '日', '日/式', '往復', 't', 'L', '人工', '台', '個', '本', 'kg', 'm', 'ｍ']:
                        unit = cell.strip()
                    else:
                        note = cell.strip()
            
            if name and (qty > 0 or amount > 0):
                items.append({
                    "project_id": project_id,
                    "category": category,
                    "work_type": name,
                    "vendor": vendor,
                    "description": f"{spec}（{note}）" if note else spec,
                    "quantity": qty,
                    "unit": unit.replace('/式', '').replace('日/式', '日'),
                    "unit_price": price,
                    "amount": amount if amount > 0 else int(qty * price)
                })
    
    # DB登録
    for item in items:
        db_detail = BudgetDetail(**item)
        db.add(db_detail)
    db.commit()
    
    return {"count": len(items), "items": items}

@app.post("/api/projects/upload-excel")
async def upload_project_excel(file: UploadFile = File(...), vendor: str = "", category: str = "外注費", db: Session = Depends(get_db)):
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    
    project_name = ""
    client = ""
    budget_items = []
    
    skip_keywords = ['直接工事費', '諸経費', '機械回送費', '小計', '合計', '端数調整', '法定福利費', '労働災害', '内訳明細書', '施工条件', '本見積', '御見積']
    
    for sheet in wb.worksheets:
        if '条件書' in sheet.title:
            continue
        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            cells = [str(c).strip() if c else "" for c in row]
            row_text = " ".join(cells)
            
            if '工事名' in row_text:
                for c in row:
                    if c and '工事名' not in str(c) and len(str(c)) > 5:
                        project_name = str(c).strip()
                        break
            
            if not client:
                for c in row:
                    if c and '御中' in str(c):
                        client = str(c).replace('御中', '').strip()
            
            if any(kw in row_text for kw in skip_keywords):
                continue
            
            nums = [c for c in row if isinstance(c, (int, float)) and c != 0]
            if len(nums) >= 2:
                name, spec, qty, unit, price, amount = "", "", 0, "式", 0, 0
                for c in row:
                    if c is None:
                        continue
                    if isinstance(c, (int, float)) and c != 0:
                        if qty == 0 and c < 100000:
                            qty = float(c)
                        elif price == 0 and c >= 100:
                            price = int(c)
                        elif amount == 0:
                            amount = int(c)
                    elif isinstance(c, str) and c.strip():
                        cs = c.strip()
                        if cs in ['m2', 'm3', '式', '日', '往復', 't', 'L', '人工', '台']:
                            unit = cs
                        elif not name and len(cs) > 2:
                            name = cs
                        elif name and not spec:
                            spec = cs
                
                if name and (qty > 0 or amount > 0):
                    if amount == 0:
                        amount = int(qty * price)
                    budget_items.append({"category": category, "work_type": name, "vendor": vendor, "description": spec, "quantity": qty, "unit": unit, "unit_price": price, "amount": amount})
    
    new_project = Project(code=str(1000 + db.query(Project).count() + 1), name=project_name or file.filename.replace('.xlsx', ''), client=client, status="見積中", order_type="一次請", prefecture="", probability="見込み有", order_amount=0, budget_amount=0, tax_rate=0.1, period="", sales_person="", site_person="")
    db.add(new_project)
    db.commit()
    db.refresh(new_project)
    
    total = 0
    for item in budget_items:
        item["project_id"] = new_project.id
        db.add(BudgetDetail(**item))
        total += item.get("amount", 0)
    
    new_project.budget_amount = total
    db.commit()
    
    return {"project_id": new_project.id, "name": new_project.name, "count": len(budget_items), "total": total}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

# ============================================
# 工種管理API（60社システム互換）
# ============================================

from models import ProjectWorkType, WorkTypeDetail

# 案件工種一覧取得
@app.get("/api/projects/{project_id}/work-types")
def get_project_work_types(project_id: int, db: Session = Depends(get_db)):
    work_types = db.query(ProjectWorkType).filter(ProjectWorkType.project_id == project_id).order_by(ProjectWorkType.seq).all()
    return work_types

# 工種作成
@app.post("/api/projects/{project_id}/work-types")
def create_project_work_type(project_id: int, data: dict, db: Session = Depends(get_db)):
    work_type = ProjectWorkType(project_id=project_id, **data)
    db.add(work_type)
    db.commit()
    db.refresh(work_type)
    return work_type

# 工種更新
@app.put("/api/work-types/{work_type_id}")
def update_project_work_type(work_type_id: int, data: dict, db: Session = Depends(get_db)):
    work_type = db.query(ProjectWorkType).filter(ProjectWorkType.id == work_type_id).first()
    if not work_type:
        raise HTTPException(status_code=404, detail="WorkType not found")
    for key, value in data.items():
        setattr(work_type, key, value)
    db.commit()
    return work_type

# 工種削除
@app.delete("/api/work-types/{work_type_id}")
def delete_project_work_type(work_type_id: int, db: Session = Depends(get_db)):
    work_type = db.query(ProjectWorkType).filter(ProjectWorkType.id == work_type_id).first()
    if work_type:
        # 明細も削除
        db.query(WorkTypeDetail).filter(WorkTypeDetail.work_type_id == work_type_id).delete()
        db.delete(work_type)
        db.commit()
    return {"ok": True}

# 工種明細一覧取得
@app.get("/api/work-types/{work_type_id}/details")
def get_work_type_details(work_type_id: int, db: Session = Depends(get_db)):
    details = db.query(WorkTypeDetail).filter(WorkTypeDetail.work_type_id == work_type_id).order_by(WorkTypeDetail.seq).all()
    return details

# 工種明細作成
@app.post("/api/work-types/{work_type_id}/details")
def create_work_type_detail(work_type_id: int, data: dict, db: Session = Depends(get_db)):
    detail = WorkTypeDetail(work_type_id=work_type_id, **data)
    db.add(detail)
    db.commit()
    db.refresh(detail)
    # 工種の予算金額を再計算
    recalc_work_type_budget(work_type_id, db)
    return detail

# 工種明細更新
@app.put("/api/work-type-details/{detail_id}")
def update_work_type_detail(detail_id: int, data: dict, db: Session = Depends(get_db)):
    detail = db.query(WorkTypeDetail).filter(WorkTypeDetail.id == detail_id).first()
    if not detail:
        raise HTTPException(status_code=404, detail="Detail not found")
    for key, value in data.items():
        setattr(detail, key, value)
    db.commit()
    # 工種の予算金額を再計算
    recalc_work_type_budget(detail.work_type_id, db)
    return detail

# 工種明細削除
@app.delete("/api/work-type-details/{detail_id}")
def delete_work_type_detail(detail_id: int, db: Session = Depends(get_db)):
    detail = db.query(WorkTypeDetail).filter(WorkTypeDetail.id == detail_id).first()
    if detail:
        work_type_id = detail.work_type_id
        db.delete(detail)
        db.commit()
        # 工種の予算金額を再計算
        recalc_work_type_budget(work_type_id, db)
    return {"ok": True}

# 工種の予算金額を再計算
def recalc_work_type_budget(work_type_id: int, db: Session):
    details = db.query(WorkTypeDetail).filter(WorkTypeDetail.work_type_id == work_type_id).all()
    total = sum(d.budget_amount or 0 for d in details)
    work_type = db.query(ProjectWorkType).filter(ProjectWorkType.id == work_type_id).first()
    if work_type:
        work_type.budget_amount = total
        # 見積金額 = 予算金額 × 掛率
        work_type.estimate_amount = int(total * (work_type.rate or 1.0))
        db.commit()


# ========== 出来高調書 (Monthly Progress) ==========

# 支払予定日を計算するヘルパー関数
def calc_payment_date(base_date: date, closing_day: int, payment_day: int, month_offset: int) -> date:
    """締め日・支払日から支払予定日を計算"""
    # 基準日が締め日を過ぎているか判定
    if base_date.day > closing_day:
        # 翌月締め
        payment_month = base_date + relativedelta(months=month_offset + 1)
    else:
        # 当月締め
        payment_month = base_date + relativedelta(months=month_offset)
    # 支払日を設定（月末を超えないよう調整）
    import calendar
    last_day = calendar.monthrange(payment_month.year, payment_month.month)[1]
    actual_payment_day = min(payment_day, last_day)
    return date(payment_month.year, payment_month.month, actual_payment_day)


@app.get("/api/progress/project/{project_id}")
def get_project_progress(project_id: int, db: Session = Depends(get_db)):
    progress = db.query(MonthlyProgress).filter(
        MonthlyProgress.project_id == project_id
    ).order_by(MonthlyProgress.year_month).all()
    return progress

@app.post("/api/progress/")
def create_progress(data: MonthlyProgressCreate, db: Session = Depends(get_db)):
    # 既存のデータがあれば更新
    existing = db.query(MonthlyProgress).filter(
        MonthlyProgress.project_id == data.project_id,
        MonthlyProgress.year_month == data.year_month
    ).first()

    if existing:
        for key, value in data.dict().items():
            setattr(existing, key, value)
        db.commit()
        progress = existing
    else:
        progress = MonthlyProgress(**data.dict())
        db.add(progress)
        db.commit()
        db.refresh(progress)

    # 自動連携: 入金予定を作成
    project = db.query(Project).filter(Project.id == data.project_id).first()
    if project and data.progress_amount > 0:
        # 元請け情報を取得
        client = db.query(Client).filter(Client.name == project.client).first()
        closing_day = client.closing_day if client else 25
        payment_day = client.payment_day if client else 25
        month_offset = client.payment_month_offset if client else 1

        # 出来高の月から支払予定日を計算
        year, month = map(int, data.year_month.split('-'))
        base_date = date(year, month, 1)
        expected_date = calc_payment_date(base_date, closing_day, payment_day, month_offset)

        # 既存の入金予定を確認
        existing_receivable = db.query(Receivable).filter(
            Receivable.progress_id == progress.id
        ).first()

        if existing_receivable:
            existing_receivable.amount = data.progress_amount
            existing_receivable.expected_date = expected_date
        else:
            receivable = Receivable(
                project_id=data.project_id,
                progress_id=progress.id,
                client_name=project.client or "未設定",
                description=f"{data.year_month} 出来高",
                amount=data.progress_amount,
                expected_date=expected_date,
                status="予定"
            )
            db.add(receivable)
        db.commit()

    return progress

@app.put("/api/progress/{progress_id}")
def update_progress(progress_id: int, data: dict, db: Session = Depends(get_db)):
    progress = db.query(MonthlyProgress).filter(MonthlyProgress.id == progress_id).first()
    if not progress:
        raise HTTPException(status_code=404, detail="Progress not found")
    for key, value in data.items():
        setattr(progress, key, value)
    db.commit()
    return progress

@app.delete("/api/progress/{progress_id}")
def delete_progress(progress_id: int, db: Session = Depends(get_db)):
    progress = db.query(MonthlyProgress).filter(MonthlyProgress.id == progress_id).first()
    if progress:
        # 関連する入金予定も削除
        db.query(Receivable).filter(Receivable.progress_id == progress_id).delete()
        db.delete(progress)
        db.commit()
    return {"ok": True}


# ========== 入金予定 (Receivables) ==========

@app.get("/api/receivables/")
def get_all_receivables(
    year_month: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(Receivable)
    if year_month:
        # 指定月の入金予定を取得
        year, month = map(int, year_month.split('-'))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        query = query.filter(Receivable.expected_date >= start_date, Receivable.expected_date < end_date)
    if status:
        query = query.filter(Receivable.status == status)
    return query.order_by(Receivable.expected_date).all()

@app.get("/api/receivables/project/{project_id}")
def get_project_receivables(project_id: int, db: Session = Depends(get_db)):
    return db.query(Receivable).filter(Receivable.project_id == project_id).order_by(Receivable.expected_date).all()

@app.post("/api/receivables/")
def create_receivable(data: ReceivableCreate, db: Session = Depends(get_db)):
    receivable = Receivable(**data.dict())
    db.add(receivable)
    db.commit()
    db.refresh(receivable)
    return receivable

@app.put("/api/receivables/{receivable_id}")
def update_receivable(receivable_id: int, data: dict, db: Session = Depends(get_db)):
    receivable = db.query(Receivable).filter(Receivable.id == receivable_id).first()
    if not receivable:
        raise HTTPException(status_code=404, detail="Receivable not found")
    for key, value in data.items():
        if key in ['expected_date', 'actual_date', 'billing_date'] and value:
            value = datetime.strptime(value, '%Y-%m-%d').date() if isinstance(value, str) else value
        setattr(receivable, key, value)
    db.commit()
    return receivable

@app.delete("/api/receivables/{receivable_id}")
def delete_receivable(receivable_id: int, db: Session = Depends(get_db)):
    receivable = db.query(Receivable).filter(Receivable.id == receivable_id).first()
    if receivable:
        db.delete(receivable)
        db.commit()
    return {"ok": True}


# ========== 支払予定 (Payables) ==========

@app.get("/api/payables/")
def get_all_payables(
    year_month: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(Payable)
    if year_month:
        year, month = map(int, year_month.split('-'))
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, month + 1, 1)
        query = query.filter(Payable.expected_date >= start_date, Payable.expected_date < end_date)
    if status:
        query = query.filter(Payable.status == status)
    return query.order_by(Payable.expected_date).all()

@app.get("/api/payables/project/{project_id}")
def get_project_payables(project_id: int, db: Session = Depends(get_db)):
    return db.query(Payable).filter(Payable.project_id == project_id).order_by(Payable.expected_date).all()

@app.post("/api/payables/")
def create_payable(data: PayableCreate, db: Session = Depends(get_db)):
    payable = Payable(**data.dict())
    db.add(payable)
    db.commit()
    db.refresh(payable)
    return payable

@app.put("/api/payables/{payable_id}")
def update_payable(payable_id: int, data: dict, db: Session = Depends(get_db)):
    payable = db.query(Payable).filter(Payable.id == payable_id).first()
    if not payable:
        raise HTTPException(status_code=404, detail="Payable not found")
    for key, value in data.items():
        if key in ['expected_date', 'actual_date', 'invoice_date'] and value:
            value = datetime.strptime(value, '%Y-%m-%d').date() if isinstance(value, str) else value
        setattr(payable, key, value)
    db.commit()
    return payable

@app.delete("/api/payables/{payable_id}")
def delete_payable(payable_id: int, db: Session = Depends(get_db)):
    payable = db.query(Payable).filter(Payable.id == payable_id).first()
    if payable:
        db.delete(payable)
        db.commit()
    return {"ok": True}


# ========== キャッシュフロー ==========

@app.get("/api/cashflow/")
def get_cashflow(year_month: Optional[str] = None, db: Session = Depends(get_db)):
    """月別のキャッシュフローを取得"""
    if not year_month:
        year_month = datetime.now().strftime('%Y-%m')

    year, month = map(int, year_month.split('-'))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    # 入金予定
    receivables = db.query(Receivable).filter(
        Receivable.expected_date >= start_date,
        Receivable.expected_date < end_date
    ).all()

    # 支払予定
    payables = db.query(Payable).filter(
        Payable.expected_date >= start_date,
        Payable.expected_date < end_date
    ).all()

    total_receivable = sum(r.amount for r in receivables)
    total_payable = sum(p.amount for p in payables)

    return {
        "year_month": year_month,
        "receivables": [{
            "id": r.id,
            "project_id": r.project_id,
            "client_name": r.client_name,
            "description": r.description,
            "amount": r.amount,
            "expected_date": r.expected_date.isoformat() if r.expected_date else None,
            "status": r.status
        } for r in receivables],
        "payables": [{
            "id": p.id,
            "project_id": p.project_id,
            "vendor_name": p.vendor_name,
            "category": p.category,
            "description": p.description,
            "amount": p.amount,
            "expected_date": p.expected_date.isoformat() if p.expected_date else None,
            "status": p.status
        } for p in payables],
        "total_receivable": total_receivable,
        "total_payable": total_payable,
        "net_cashflow": total_receivable - total_payable
    }


# ========== 原価登録時の自動支払予定作成 ==========
# 既存の原価登録エンドポイントを拡張する必要がある場合はここに追加


# ========== ジオコーディングAPI ==========
@app.get("/api/geocode")
async def geocode_address(address: str):
    """住所から緯度経度を取得（国土地理院API使用）"""
    try:
        async with httpx.AsyncClient() as client:
            # 国土地理院のジオコーディングAPI
            url = f"https://msearch.gsi.go.jp/address-search/AddressSearch?q={address}"
            response = await client.get(url)

            if response.status_code == 200:
                data = response.json()
                if data and len(data) > 0:
                    # 最初の結果を使用
                    result = data[0]
                    coordinates = result.get("geometry", {}).get("coordinates", [])
                    if len(coordinates) >= 2:
                        return {
                            "success": True,
                            "latitude": coordinates[1],  # 緯度
                            "longitude": coordinates[0],  # 経度
                            "address": result.get("properties", {}).get("title", address)
                        }

            return {"success": False, "error": "住所が見つかりませんでした"}
    except Exception as e:
        return {"success": False, "error": str(e)}


# ========== 天気予報API ==========
@app.get("/api/weather")
async def get_weather(lat: float, lon: float):
    """緯度経度から14日間の天気予報を取得（Open-Meteo API使用）"""
    try:
        async with httpx.AsyncClient() as client:
            # Open-Meteo APIを使用（無料・APIキー不要）
            url = (
                f"https://api.open-meteo.com/v1/forecast?"
                f"latitude={lat}&longitude={lon}"
                f"&daily=weather_code,temperature_2m_max,temperature_2m_min,precipitation_probability_max"
                f"&timezone=Asia/Tokyo"
                f"&forecast_days=14"
            )
            response = await client.get(url)

            if response.status_code == 200:
                data = response.json()
                daily = data.get("daily", {})

                # 天気コードから天気情報に変換
                weather_codes = {
                    0: {"icon": "☀️", "text": "快晴"},
                    1: {"icon": "🌤️", "text": "晴れ"},
                    2: {"icon": "⛅", "text": "薄曇り"},
                    3: {"icon": "☁️", "text": "曇り"},
                    45: {"icon": "🌫️", "text": "霧"},
                    48: {"icon": "🌫️", "text": "霧氷"},
                    51: {"icon": "🌧️", "text": "小雨"},
                    53: {"icon": "🌧️", "text": "雨"},
                    55: {"icon": "🌧️", "text": "強い雨"},
                    61: {"icon": "🌧️", "text": "弱い雨"},
                    63: {"icon": "🌧️", "text": "雨"},
                    65: {"icon": "🌧️", "text": "強い雨"},
                    71: {"icon": "🌨️", "text": "小雪"},
                    73: {"icon": "🌨️", "text": "雪"},
                    75: {"icon": "🌨️", "text": "大雪"},
                    80: {"icon": "🌦️", "text": "にわか雨"},
                    81: {"icon": "🌦️", "text": "にわか雨"},
                    82: {"icon": "🌧️", "text": "激しいにわか雨"},
                    95: {"icon": "⛈️", "text": "雷雨"},
                    96: {"icon": "⛈️", "text": "雷雨・雹"},
                    99: {"icon": "⛈️", "text": "激しい雷雨"},
                }

                forecasts = []
                dates = daily.get("time", [])
                codes = daily.get("weather_code", [])
                temp_max = daily.get("temperature_2m_max", [])
                temp_min = daily.get("temperature_2m_min", [])
                precip_prob = daily.get("precipitation_probability_max", [])

                for i in range(len(dates)):
                    code = codes[i] if i < len(codes) else 0
                    weather = weather_codes.get(code, {"icon": "❓", "text": "不明"})
                    forecasts.append({
                        "date": dates[i],
                        "weather_code": code,
                        "icon": weather["icon"],
                        "text": weather["text"],
                        "temp_max": temp_max[i] if i < len(temp_max) else None,
                        "temp_min": temp_min[i] if i < len(temp_min) else None,
                        "precipitation_probability": precip_prob[i] if i < len(precip_prob) else None
                    })

                return {
                    "success": True,
                    "forecasts": forecasts
                }

            return {"success": False, "error": "天気情報を取得できませんでした"}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.get("/api/projects/with-weather")
async def get_projects_with_weather(db: Session = Depends(get_db)):
    """進行中の現場と天気情報を取得"""
    # 進行中の現場のみ取得
    projects = db.query(Project).filter(
        Project.status.in_(["施工中", "受注確定"])
    ).all()

    result = []
    async with httpx.AsyncClient() as client:
        for project in projects:
            project_data = {
                "id": project.id,
                "code": project.code,
                "name": project.name,
                "client": project.client,
                "status": project.status,
                "address": project.address,
                "latitude": project.latitude,
                "longitude": project.longitude,
                "weather": None
            }

            # 位置情報がある場合は天気を取得
            if project.latitude and project.longitude:
                try:
                    url = (
                        f"https://api.open-meteo.com/v1/forecast?"
                        f"latitude={project.latitude}&longitude={project.longitude}"
                        f"&daily=weather_code,temperature_2m_max,temperature_2m_min,precipitation_probability_max"
                        f"&timezone=Asia/Tokyo"
                        f"&forecast_days=14"
                    )
                    response = await client.get(url)
                    if response.status_code == 200:
                        data = response.json()
                        daily = data.get("daily", {})

                        weather_codes = {
                            0: {"icon": "☀️", "text": "快晴"},
                            1: {"icon": "🌤️", "text": "晴れ"},
                            2: {"icon": "⛅", "text": "薄曇り"},
                            3: {"icon": "☁️", "text": "曇り"},
                            45: {"icon": "🌫️", "text": "霧"},
                            48: {"icon": "🌫️", "text": "霧氷"},
                            51: {"icon": "🌧️", "text": "小雨"},
                            53: {"icon": "🌧️", "text": "雨"},
                            55: {"icon": "🌧️", "text": "強い雨"},
                            61: {"icon": "🌧️", "text": "弱い雨"},
                            63: {"icon": "🌧️", "text": "雨"},
                            65: {"icon": "🌧️", "text": "強い雨"},
                            71: {"icon": "🌨️", "text": "小雪"},
                            73: {"icon": "🌨️", "text": "雪"},
                            75: {"icon": "🌨️", "text": "大雪"},
                            80: {"icon": "🌦️", "text": "にわか雨"},
                            81: {"icon": "🌦️", "text": "にわか雨"},
                            82: {"icon": "🌧️", "text": "激しいにわか雨"},
                            95: {"icon": "⛈️", "text": "雷雨"},
                            96: {"icon": "⛈️", "text": "雷雨・雹"},
                            99: {"icon": "⛈️", "text": "激しい雷雨"},
                        }

                        forecasts = []
                        dates = daily.get("time", [])
                        codes = daily.get("weather_code", [])
                        temp_max = daily.get("temperature_2m_max", [])
                        temp_min = daily.get("temperature_2m_min", [])
                        precip_prob = daily.get("precipitation_probability_max", [])

                        for i in range(len(dates)):
                            code = codes[i] if i < len(codes) else 0
                            weather = weather_codes.get(code, {"icon": "❓", "text": "不明"})
                            forecasts.append({
                                "date": dates[i],
                                "weather_code": code,
                                "icon": weather["icon"],
                                "text": weather["text"],
                                "temp_max": temp_max[i] if i < len(temp_max) else None,
                                "temp_min": temp_min[i] if i < len(temp_min) else None,
                                "precipitation_probability": precip_prob[i] if i < len(precip_prob) else None
                            })

                        project_data["weather"] = forecasts
                except Exception:
                    pass

            result.append(project_data)

    return result


# ========== Estimates API ==========
@app.get("/api/estimates/project/{project_id}")
def get_estimates_by_project(project_id: int, db: Session = Depends(get_db)):
    estimates = db.query(Estimate).filter(Estimate.project_id == project_id).all()
    return [{"id": e.id, "project_id": e.project_id, "work_type": e.work_type,
             "description": e.description, "quantity": e.quantity, "unit": e.unit,
             "unit_price": e.unit_price, "amount": e.amount, "rate": e.rate} for e in estimates]

@app.post("/api/estimates/")
def create_estimate(data: EstimateCreate, db: Session = Depends(get_db)):
    estimate = Estimate(**data.model_dump())
    db.add(estimate)
    db.commit()
    db.refresh(estimate)
    return estimate

@app.put("/api/estimates/{estimate_id}")
def update_estimate(estimate_id: int, data: EstimateCreate, db: Session = Depends(get_db)):
    estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")
    for key, value in data.model_dump().items():
        setattr(estimate, key, value)
    db.commit()
    return estimate

@app.delete("/api/estimates/{estimate_id}")
def delete_estimate(estimate_id: int, db: Session = Depends(get_db)):
    estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
    if estimate:
        db.delete(estimate)
        db.commit()
    return {"message": "deleted"}


# ========== Budgets API ==========
@app.get("/api/budgets/project/{project_id}")
def get_budgets_by_project(project_id: int, db: Session = Depends(get_db)):
    budgets = db.query(Budget).filter(Budget.project_id == project_id).all()
    return [{"id": b.id, "project_id": b.project_id, "work_type": b.work_type,
             "category": b.category, "vendor": b.vendor, "description": b.description,
             "quantity": b.quantity, "unit": b.unit, "unit_price": b.unit_price,
             "amount": b.amount} for b in budgets]

@app.post("/api/budgets/")
def create_budget(data: BudgetCreate, db: Session = Depends(get_db)):
    budget = Budget(**data.model_dump())
    db.add(budget)
    db.commit()
    db.refresh(budget)
    return budget

@app.post("/api/budgets/bulk")
def create_budgets_bulk(items: List[BudgetCreate], db: Session = Depends(get_db)):
    budgets = [Budget(**item.model_dump()) for item in items]
    db.add_all(budgets)
    db.commit()
    return {"message": f"Created {len(budgets)} budgets"}

@app.put("/api/budgets/{budget_id}")
def update_budget(budget_id: int, data: BudgetCreate, db: Session = Depends(get_db)):
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if not budget:
        raise HTTPException(status_code=404, detail="Budget not found")
    for key, value in data.model_dump().items():
        setattr(budget, key, value)
    db.commit()
    return budget

@app.delete("/api/budgets/{budget_id}")
def delete_budget(budget_id: int, db: Session = Depends(get_db)):
    budget = db.query(Budget).filter(Budget.id == budget_id).first()
    if budget:
        db.delete(budget)
        db.commit()
    return {"message": "deleted"}


# ========== Costs by Project API ==========
@app.get("/api/costs/project/{project_id}")
def get_costs_by_project(project_id: int, db: Session = Depends(get_db)):
    costs = db.query(Cost).filter(Cost.project_id == project_id).order_by(Cost.date.desc()).all()
    return [{"id": c.id, "project_id": c.project_id, "date": c.date.isoformat() if c.date else None,
             "category": c.category, "work_type": c.work_type, "vendor": c.vendor,
             "description": c.description, "quantity": c.quantity, "unit": c.unit,
             "unit_price": c.unit_price, "amount": c.amount} for c in costs]


# ========== Expenses API ==========
@app.get("/api/expenses/")
def get_expenses(project_id: Optional[int] = None, status: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Expense)
    if project_id:
        query = query.filter(Expense.project_id == project_id)
    if status:
        query = query.filter(Expense.status == status)
    expenses = query.order_by(Expense.expense_date.desc()).all()
    return expenses

@app.get("/api/expenses/{expense_id}")
def get_expense(expense_id: int, db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    return expense

@app.post("/api/expenses/")
def create_expense(data: ExpenseCreate, db: Session = Depends(get_db)):
    expense = Expense(**data.model_dump())
    db.add(expense)
    db.commit()
    db.refresh(expense)
    # Create approval record
    approval = Approval(type="expense", reference_id=expense.id, requested_by="user")
    db.add(approval)
    db.commit()
    return expense

@app.put("/api/expenses/{expense_id}")
def update_expense(expense_id: int, data: ExpenseCreate, db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    for key, value in data.model_dump().items():
        setattr(expense, key, value)
    db.commit()
    return expense

@app.delete("/api/expenses/{expense_id}")
def delete_expense(expense_id: int, db: Session = Depends(get_db)):
    expense = db.query(Expense).filter(Expense.id == expense_id).first()
    if expense:
        db.delete(expense)
        db.commit()
    return {"message": "deleted"}


# ========== Expense Categories API ==========
@app.get("/api/expense-categories/")
def get_expense_categories(db: Session = Depends(get_db)):
    categories = db.query(ExpenseCategory).filter(ExpenseCategory.is_active == True).order_by(ExpenseCategory.sort_order).all()
    return categories

@app.post("/api/expense-categories/")
def create_expense_category(name: str, icon: Optional[str] = None, is_fuel: Optional[bool] = False, db: Session = Depends(get_db)):
    category = ExpenseCategory(name=name, icon=icon, is_fuel=is_fuel)
    db.add(category)
    db.commit()
    db.refresh(category)
    return category


# ========== Fuel Prices API ==========
@app.get("/api/fuel-prices/")
def get_fuel_prices(year_month: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(FuelPrice)
    if year_month:
        query = query.filter(FuelPrice.year_month == year_month)
    return query.order_by(FuelPrice.year_month.desc()).all()

@app.get("/api/fuel-prices/latest")
def get_latest_fuel_price(db: Session = Depends(get_db)):
    """最新の燃料単価を取得"""
    price = db.query(FuelPrice).order_by(FuelPrice.year_month.desc()).first()
    if not price:
        # デフォルト値を返す
        return {"year_month": datetime.now().strftime("%Y-%m"), "fuel_type": "regular", "price": 170}
    return {"year_month": price.year_month, "fuel_type": price.fuel_type, "price": price.price_per_liter}

@app.post("/api/fuel-prices/")
def create_fuel_price(year_month: str, fuel_type: str, price_per_liter: float, db: Session = Depends(get_db)):
    price = FuelPrice(year_month=year_month, fuel_type=fuel_type, price_per_liter=price_per_liter)
    db.add(price)
    db.commit()
    db.refresh(price)
    return price


# ========== Billings API ==========
@app.get("/api/billings/")
def get_billings(project_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Billing)
    if project_id:
        query = query.filter(Billing.project_id == project_id)
    return query.order_by(Billing.date.desc()).all()

@app.post("/api/billings/")
def create_billing(data: BillingCreate, db: Session = Depends(get_db)):
    billing = Billing(**data.model_dump())
    db.add(billing)
    db.commit()
    db.refresh(billing)
    return billing

@app.put("/api/billings/{billing_id}")
def update_billing(billing_id: int, data: BillingCreate, db: Session = Depends(get_db)):
    billing = db.query(Billing).filter(Billing.id == billing_id).first()
    if not billing:
        raise HTTPException(status_code=404, detail="Billing not found")
    for key, value in data.model_dump().items():
        setattr(billing, key, value)
    db.commit()
    return billing

@app.delete("/api/billings/{billing_id}")
def delete_billing(billing_id: int, db: Session = Depends(get_db)):
    billing = db.query(Billing).filter(Billing.id == billing_id).first()
    if billing:
        db.delete(billing)
        db.commit()
    return {"message": "deleted"}


# ========== Approvals API ==========
@app.get("/api/approvals/")
def get_approvals(status: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Approval)
    if status:
        query = query.filter(Approval.status == status)
    return query.order_by(Approval.requested_at.desc()).all()

@app.get("/api/approvals/pending")
def get_pending_approvals(db: Session = Depends(get_db)):
    approvals = db.query(Approval).filter(Approval.status == "pending").order_by(Approval.requested_at.desc()).all()
    return approvals

@app.get("/api/approvals/count")
def get_pending_count(db: Session = Depends(get_db)):
    count = db.query(Approval).filter(Approval.status == "pending").count()
    return {"count": count}

@app.put("/api/approvals/{approval_id}/approve")
def approve_approval(approval_id: int, approved_by: Optional[str] = None, db: Session = Depends(get_db)):
    approval = db.query(Approval).filter(Approval.id == approval_id).first()
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    approval.status = "approved"
    approval.approved_by = approved_by
    approval.approved_at = datetime.now()
    db.commit()
    return approval

@app.put("/api/approvals/{approval_id}/reject")
def reject_approval(approval_id: int, comment: Optional[str] = None, db: Session = Depends(get_db)):
    approval = db.query(Approval).filter(Approval.id == approval_id).first()
    if not approval:
        raise HTTPException(status_code=404, detail="Approval not found")
    approval.status = "rejected"
    approval.comment = comment
    approval.approved_at = datetime.now()
    db.commit()
    return approval


# ========== Notifications API ==========
@app.get("/api/notifications/")
def get_notifications(user_id: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Notification)
    if user_id:
        query = query.filter(Notification.user_id == user_id)
    return query.order_by(Notification.created_at.desc()).limit(50).all()

@app.get("/api/notifications/unread-count")
def get_unread_count(user_id: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Notification).filter(Notification.is_read == False)
    if user_id:
        query = query.filter(Notification.user_id == user_id)
    return {"count": query.count()}

@app.post("/api/notifications/")
def create_notification(data: NotificationCreate, db: Session = Depends(get_db)):
    notification = Notification(**data.model_dump())
    db.add(notification)
    db.commit()
    db.refresh(notification)
    return notification

@app.put("/api/notifications/{notification_id}/read")
def mark_notification_read(notification_id: int, db: Session = Depends(get_db)):
    notification = db.query(Notification).filter(Notification.id == notification_id).first()
    if notification:
        notification.is_read = True
        db.commit()
    return {"message": "marked as read"}

@app.put("/api/notifications/read-all")
def mark_all_read(user_id: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Notification).filter(Notification.is_read == False)
    if user_id:
        query = query.filter(Notification.user_id == user_id)
    query.update({"is_read": True})
    db.commit()
    return {"message": "all marked as read"}


# ========== Workers API ==========
@app.get("/api/workers/")
def get_workers(team: Optional[str] = None, is_active: Optional[bool] = True, db: Session = Depends(get_db)):
    query = db.query(Worker)
    if team:
        query = query.filter(Worker.team == team)
    if is_active is not None:
        query = query.filter(Worker.is_active == is_active)
    return query.order_by(Worker.name).all()

@app.post("/api/workers/")
def create_worker(data: WorkerCreate, db: Session = Depends(get_db)):
    worker = Worker(**data.model_dump())
    db.add(worker)
    db.commit()
    db.refresh(worker)
    return worker

@app.put("/api/workers/{worker_id}")
def update_worker(worker_id: int, data: WorkerCreate, db: Session = Depends(get_db)):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if not worker:
        raise HTTPException(status_code=404, detail="Worker not found")
    for key, value in data.model_dump().items():
        setattr(worker, key, value)
    db.commit()
    return worker

@app.delete("/api/workers/{worker_id}")
def delete_worker(worker_id: int, db: Session = Depends(get_db)):
    worker = db.query(Worker).filter(Worker.id == worker_id).first()
    if worker:
        worker.is_active = False
        db.commit()
    return {"message": "deactivated"}


# ========== Assignments API ==========
@app.get("/api/assignments/")
def get_assignments(date_str: Optional[str] = None, week: Optional[str] = None, project_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Assignment)
    if date_str:
        query = query.filter(Assignment.date == date_str)
    if week:
        week_start = datetime.strptime(week, "%Y-%m-%d").date()
        week_end = week_start + relativedelta(days=6)
        query = query.filter(Assignment.date >= week_start, Assignment.date <= week_end)
    if project_id:
        query = query.filter(Assignment.project_id == project_id)
    assignments = query.order_by(Assignment.date).all()
    result = []
    for a in assignments:
        result.append({
            "id": a.id,
            "date": a.date.isoformat() if a.date else None,
            "project_id": a.project_id,
            "worker_id": a.worker_id,
            "start_time": a.start_time,
            "end_time": a.end_time,
            "note": a.note
        })
    return result

@app.post("/api/assignments/")
def create_assignment(data: AssignmentCreate, db: Session = Depends(get_db)):
    assignment = Assignment(**data.model_dump())
    db.add(assignment)
    db.commit()
    db.refresh(assignment)
    return assignment

@app.put("/api/assignments/{assignment_id}")
def update_assignment(assignment_id: int, data: AssignmentCreate, db: Session = Depends(get_db)):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="Assignment not found")
    for key, value in data.model_dump().items():
        setattr(assignment, key, value)
    db.commit()
    return assignment

@app.delete("/api/assignments/{assignment_id}")
def delete_assignment(assignment_id: int, db: Session = Depends(get_db)):
    assignment = db.query(Assignment).filter(Assignment.id == assignment_id).first()
    if assignment:
        db.delete(assignment)
        db.commit()
    return {"message": "deleted"}


# ========== KY Reports API ==========
@app.get("/api/ky-reports/")
def get_ky_reports(project_id: Optional[int] = None, date_str: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(KYReport)
    if project_id:
        query = query.filter(KYReport.project_id == project_id)
    if date_str:
        query = query.filter(KYReport.date == date_str)
    return query.order_by(KYReport.date.desc()).all()

@app.get("/api/ky-reports/{report_id}")
def get_ky_report(report_id: int, db: Session = Depends(get_db)):
    report = db.query(KYReport).filter(KYReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="KY Report not found")
    return report

@app.post("/api/ky-reports/")
def create_ky_report(data: KYReportCreate, db: Session = Depends(get_db)):
    report = KYReport(**data.model_dump())
    db.add(report)
    db.commit()
    db.refresh(report)
    return report

@app.put("/api/ky-reports/{report_id}")
def update_ky_report(report_id: int, data: KYReportCreate, db: Session = Depends(get_db)):
    report = db.query(KYReport).filter(KYReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="KY Report not found")
    for key, value in data.model_dump().items():
        setattr(report, key, value)
    db.commit()
    return report

@app.post("/api/ky-reports/{report_id}/sign")
def sign_ky_report(report_id: int, worker_id: int, db: Session = Depends(get_db)):
    signature = KYSignature(ky_report_id=report_id, worker_id=worker_id, signed_at=datetime.now())
    db.add(signature)
    db.commit()
    return {"message": "signed"}


# ========== Inventory API ==========
@app.get("/api/inventory/")
def get_inventory(category: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(InventoryItem)
    if category:
        query = query.filter(InventoryItem.category == category)
    return query.order_by(InventoryItem.name).all()

@app.get("/api/inventory/alerts")
def get_inventory_alerts(db: Session = Depends(get_db)):
    items = db.query(InventoryItem).filter(InventoryItem.quantity <= InventoryItem.min_quantity).all()
    return items

@app.post("/api/inventory/")
def create_inventory_item(data: InventoryItemCreate, db: Session = Depends(get_db)):
    item = InventoryItem(**data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item

@app.put("/api/inventory/{item_id}")
def update_inventory_item(item_id: int, data: InventoryItemCreate, db: Session = Depends(get_db)):
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    for key, value in data.model_dump().items():
        setattr(item, key, value)
    db.commit()
    return item

@app.delete("/api/inventory/{item_id}")
def delete_inventory_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if item:
        db.delete(item)
        db.commit()
    return {"message": "deleted"}

@app.post("/api/inventory/{item_id}/in")
def inventory_in(item_id: int, data: InventoryTransactionCreate, db: Session = Depends(get_db)):
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    item.quantity += data.quantity
    transaction = InventoryTransaction(item_id=item_id, type="in", quantity=data.quantity,
                                        project_id=data.project_id, date=data.date or date.today(), note=data.note)
    db.add(transaction)
    db.commit()
    return item

@app.post("/api/inventory/{item_id}/out")
def inventory_out(item_id: int, data: InventoryTransactionCreate, db: Session = Depends(get_db)):
    item = db.query(InventoryItem).filter(InventoryItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    item.quantity -= data.quantity
    transaction = InventoryTransaction(item_id=item_id, type="out", quantity=data.quantity,
                                        project_id=data.project_id, date=data.date or date.today(), note=data.note)
    db.add(transaction)
    db.commit()
    return item


# ========== Vehicles API ==========
@app.get("/api/vehicles/")
def get_vehicles(db: Session = Depends(get_db)):
    return db.query(Vehicle).order_by(Vehicle.name).all()

@app.get("/api/vehicles/alerts")
def get_vehicle_alerts(db: Session = Depends(get_db)):
    today = date.today()
    alert_date = today + relativedelta(days=30)
    vehicles = db.query(Vehicle).filter(
        (Vehicle.inspection_date <= alert_date) | (Vehicle.insurance_date <= alert_date)
    ).all()
    return vehicles

@app.post("/api/vehicles/")
def create_vehicle(data: VehicleCreate, db: Session = Depends(get_db)):
    vehicle = Vehicle(**data.model_dump())
    db.add(vehicle)
    db.commit()
    db.refresh(vehicle)
    return vehicle

@app.put("/api/vehicles/{vehicle_id}")
def update_vehicle(vehicle_id: int, data: VehicleCreate, db: Session = Depends(get_db)):
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehicle not found")
    for key, value in data.model_dump().items():
        setattr(vehicle, key, value)
    db.commit()
    return vehicle

@app.delete("/api/vehicles/{vehicle_id}")
def delete_vehicle(vehicle_id: int, db: Session = Depends(get_db)):
    vehicle = db.query(Vehicle).filter(Vehicle.id == vehicle_id).first()
    if vehicle:
        db.delete(vehicle)
        db.commit()
    return {"message": "deleted"}


# ========== Vehicle Logs API ==========
@app.get("/api/vehicle-logs/")
def get_vehicle_logs(vehicle_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(VehicleLog)
    if vehicle_id:
        query = query.filter(VehicleLog.vehicle_id == vehicle_id)
    return query.order_by(VehicleLog.date.desc()).all()

@app.post("/api/vehicle-logs/")
def create_vehicle_log(data: VehicleLogCreate, db: Session = Depends(get_db)):
    log = VehicleLog(**data.model_dump())
    db.add(log)
    db.commit()
    db.refresh(log)
    return log


# ========== Schedules API ==========
@app.get("/api/schedules/")
def get_schedules(db: Session = Depends(get_db)):
    schedules = db.query(Schedule).all()
    result = []
    for s in schedules:
        project = db.query(Project).filter(Project.id == s.project_id).first()
        result.append({
            "id": s.id,
            "project_id": s.project_id,
            "project_name": project.name if project else None,
            "start_date": s.start_date.isoformat() if s.start_date else None,
            "end_date": s.end_date.isoformat() if s.end_date else None,
            "progress_rate": s.progress_rate,
            "color": s.color
        })
    return result

@app.post("/api/schedules/")
def create_schedule(data: ScheduleCreate, db: Session = Depends(get_db)):
    schedule = Schedule(**data.model_dump())
    db.add(schedule)
    db.commit()
    db.refresh(schedule)
    return schedule

@app.put("/api/schedules/{schedule_id}")
def update_schedule(schedule_id: int, data: ScheduleCreate, db: Session = Depends(get_db)):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    for key, value in data.model_dump().items():
        setattr(schedule, key, value)
    db.commit()
    return schedule

@app.delete("/api/schedules/{schedule_id}")
def delete_schedule(schedule_id: int, db: Session = Depends(get_db)):
    schedule = db.query(Schedule).filter(Schedule.id == schedule_id).first()
    if schedule:
        db.delete(schedule)
        db.commit()
    return {"message": "deleted"}


# ========== OCR API ==========
@app.post("/api/ocr/schedule")
async def ocr_schedule(file: UploadFile = File(...)):
    """
    工程表画像をOCRで読み取り、構造化データを返す
    現在はダミーデータを返すが、将来的にはOCRサービスに接続予定
    """
    import re
    from datetime import datetime

    # ファイル情報を取得
    filename = file.filename or ""
    content_type = file.content_type or ""

    # ファイル名から工事名を推測
    project_name = "新規工事"
    if filename:
        # 拡張子を除去
        name_without_ext = re.sub(r'\.[^.]+$', '', filename)
        # 一般的な接尾辞を除去
        cleaned_name = re.sub(r'(工程表|schedule|_|-)$', '', name_without_ext, flags=re.IGNORECASE)
        if cleaned_name and len(cleaned_name) > 0:
            project_name = cleaned_name

    # 現在の年度を基準にデフォルト期間を設定
    now = datetime.now()
    current_year = now.year

    # 今が1〜3月なら前年度、それ以外は今年度
    if now.month <= 3:
        fiscal_year = current_year - 1
    else:
        fiscal_year = current_year

    # ダミーデータを返す（将来的にはOCR結果を返す）
    return {
        "project_name": project_name,
        "start_date": f"{fiscal_year}-04-01",
        "end_date": f"{fiscal_year + 1}-03-31",
        "progress_rate": 0,
        "color": "#3b82f6",
        "message": "OCR処理完了（デモモード）"
    }


# ========== Attendances API ==========
@app.get("/api/attendances/")
def get_attendances(worker_id: Optional[int] = None, month: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Attendance)
    if worker_id:
        query = query.filter(Attendance.worker_id == worker_id)
    if month:
        year, mon = map(int, month.split("-"))
        start_date = date(year, mon, 1)
        if mon == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, mon + 1, 1)
        query = query.filter(Attendance.date >= start_date, Attendance.date < end_date)
    return query.order_by(Attendance.date.desc()).all()

@app.get("/api/attendances/summary")
def get_attendance_summary(month: str, db: Session = Depends(get_db)):
    year, mon = map(int, month.split("-"))
    start_date = date(year, mon, 1)
    if mon == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, mon + 1, 1)

    workers = db.query(Worker).filter(Worker.is_active == True).all()
    result = []
    for worker in workers:
        attendances = db.query(Attendance).filter(
            Attendance.worker_id == worker.id,
            Attendance.date >= start_date,
            Attendance.date < end_date
        ).all()
        total_days = len(attendances)
        total_overtime = sum(a.overtime_hours or 0 for a in attendances)
        result.append({
            "worker_id": worker.id,
            "worker_name": worker.name,
            "total_days": total_days,
            "total_overtime": total_overtime
        })
    return result

@app.post("/api/attendances/")
def create_attendance(data: AttendanceCreate, db: Session = Depends(get_db)):
    attendance = Attendance(**data.model_dump())
    db.add(attendance)
    db.commit()
    db.refresh(attendance)
    return attendance

@app.put("/api/attendances/{attendance_id}")
def update_attendance(attendance_id: int, data: AttendanceCreate, db: Session = Depends(get_db)):
    attendance = db.query(Attendance).filter(Attendance.id == attendance_id).first()
    if not attendance:
        raise HTTPException(status_code=404, detail="Attendance not found")
    for key, value in data.model_dump().items():
        setattr(attendance, key, value)
    db.commit()
    return attendance

@app.post("/api/attendances/check-in")
def check_in(worker_id: int, project_id: Optional[int] = None, db: Session = Depends(get_db)):
    today = date.today()
    attendance = db.query(Attendance).filter(
        Attendance.worker_id == worker_id,
        Attendance.date == today
    ).first()
    if not attendance:
        attendance = Attendance(worker_id=worker_id, date=today, project_id=project_id, check_in=datetime.now())
        db.add(attendance)
    else:
        attendance.check_in = datetime.now()
    db.commit()
    return attendance

@app.post("/api/attendances/check-out")
def check_out(worker_id: int, db: Session = Depends(get_db)):
    today = date.today()
    attendance = db.query(Attendance).filter(
        Attendance.worker_id == worker_id,
        Attendance.date == today
    ).first()
    if attendance:
        attendance.check_out = datetime.now()
        if attendance.check_in:
            hours = (attendance.check_out - attendance.check_in).total_seconds() / 3600
            if hours > 8:
                attendance.overtime_hours = hours - 8
        db.commit()
    return attendance


# ========== Subcontractors API ==========
@app.get("/api/subcontractors/")
def get_subcontractors(category: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Subcontractor)
    if category:
        query = query.filter(Subcontractor.category == category)
    return query.order_by(Subcontractor.name).all()

@app.post("/api/subcontractors/")
def create_subcontractor(data: SubcontractorCreate, db: Session = Depends(get_db)):
    sub = Subcontractor(**data.model_dump())
    db.add(sub)
    db.commit()
    db.refresh(sub)
    return sub

@app.put("/api/subcontractors/{sub_id}")
def update_subcontractor(sub_id: int, data: SubcontractorCreate, db: Session = Depends(get_db)):
    sub = db.query(Subcontractor).filter(Subcontractor.id == sub_id).first()
    if not sub:
        raise HTTPException(status_code=404, detail="Subcontractor not found")
    for key, value in data.model_dump().items():
        setattr(sub, key, value)
    db.commit()
    return sub

@app.delete("/api/subcontractors/{sub_id}")
def delete_subcontractor(sub_id: int, db: Session = Depends(get_db)):
    sub = db.query(Subcontractor).filter(Subcontractor.id == sub_id).first()
    if sub:
        db.delete(sub)
        db.commit()
    return {"message": "deleted"}


# ========== Orders API ==========
@app.get("/api/orders/")
def get_orders(project_id: Optional[int] = None, status: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Order)
    if project_id:
        query = query.filter(Order.project_id == project_id)
    if status:
        query = query.filter(Order.status == status)
    return query.order_by(Order.order_date.desc()).all()

@app.post("/api/orders/")
def create_order(data: OrderCreate, db: Session = Depends(get_db)):
    order = Order(**data.model_dump())
    db.add(order)
    db.commit()
    db.refresh(order)
    return order

@app.put("/api/orders/{order_id}")
def update_order(order_id: int, data: OrderCreate, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    for key, value in data.model_dump().items():
        setattr(order, key, value)
    db.commit()
    return order

@app.put("/api/orders/{order_id}/status")
def update_order_status(order_id: int, status: str, db: Session = Depends(get_db)):
    order = db.query(Order).filter(Order.id == order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    order.status = status
    db.commit()
    return order


# ========== Checklists API ==========
@app.get("/api/checklists/")
def get_checklists(project_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Checklist)
    if project_id:
        query = query.filter(Checklist.project_id == project_id)
    return query.order_by(Checklist.created_at.desc()).all()

@app.post("/api/checklists/")
def create_checklist(data: ChecklistCreate, db: Session = Depends(get_db)):
    checklist = Checklist(**data.model_dump())
    db.add(checklist)
    db.commit()
    db.refresh(checklist)
    return checklist

@app.put("/api/checklists/{checklist_id}")
def update_checklist(checklist_id: int, data: ChecklistCreate, db: Session = Depends(get_db)):
    checklist = db.query(Checklist).filter(Checklist.id == checklist_id).first()
    if not checklist:
        raise HTTPException(status_code=404, detail="Checklist not found")
    for key, value in data.model_dump().items():
        setattr(checklist, key, value)
    db.commit()
    return checklist


# ========== Emergency Contacts API ==========
@app.get("/api/emergency-contacts/")
def get_emergency_contacts(db: Session = Depends(get_db)):
    return db.query(EmergencyContact).order_by(EmergencyContact.priority).all()

@app.post("/api/emergency-contacts/")
def create_emergency_contact(data: EmergencyContactCreate, db: Session = Depends(get_db)):
    contact = EmergencyContact(**data.model_dump())
    db.add(contact)
    db.commit()
    db.refresh(contact)
    return contact

@app.put("/api/emergency-contacts/{contact_id}")
def update_emergency_contact(contact_id: int, data: EmergencyContactCreate, db: Session = Depends(get_db)):
    contact = db.query(EmergencyContact).filter(EmergencyContact.id == contact_id).first()
    if not contact:
        raise HTTPException(status_code=404, detail="Emergency contact not found")
    for key, value in data.model_dump().items():
        setattr(contact, key, value)
    db.commit()
    return contact

@app.delete("/api/emergency-contacts/{contact_id}")
def delete_emergency_contact(contact_id: int, db: Session = Depends(get_db)):
    contact = db.query(EmergencyContact).filter(EmergencyContact.id == contact_id).first()
    if contact:
        db.delete(contact)
        db.commit()
    return {"message": "deleted"}


# ========== Equipment API ==========
@app.get("/api/equipment/")
def get_equipment(status: Optional[str] = None, category: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Equipment)
    if status:
        query = query.filter(Equipment.status == status)
    if category:
        query = query.filter(Equipment.category == category)
    return query.order_by(Equipment.name).all()

@app.post("/api/equipment/")
def create_equipment(data: EquipmentCreate, db: Session = Depends(get_db)):
    equipment = Equipment(**data.model_dump())
    db.add(equipment)
    db.commit()
    db.refresh(equipment)
    return equipment

@app.put("/api/equipment/{equipment_id}")
def update_equipment(equipment_id: int, data: EquipmentCreate, db: Session = Depends(get_db)):
    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="Equipment not found")
    for key, value in data.model_dump().items():
        setattr(equipment, key, value)
    db.commit()
    return equipment


# ========== Templates API ==========
@app.get("/api/templates/")
def get_templates(type: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Template)
    if type:
        query = query.filter(Template.type == type)
    return query.order_by(Template.name).all()

@app.post("/api/templates/")
def create_template(data: TemplateCreate, db: Session = Depends(get_db)):
    template = Template(**data.model_dump())
    db.add(template)
    db.commit()
    db.refresh(template)
    return template

@app.put("/api/templates/{template_id}")
def update_template(template_id: int, data: TemplateCreate, db: Session = Depends(get_db)):
    template = db.query(Template).filter(Template.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    for key, value in data.model_dump().items():
        setattr(template, key, value)
    db.commit()
    return template

@app.delete("/api/templates/{template_id}")
def delete_template(template_id: int, db: Session = Depends(get_db)):
    template = db.query(Template).filter(Template.id == template_id).first()
    if template:
        db.delete(template)
        db.commit()
    return {"message": "deleted"}


# ========== Search API ==========
@app.get("/api/search/")
def search(q: str, type: Optional[str] = None, db: Session = Depends(get_db)):
    results = []
    search_term = f"%{q}%"

    if not type or type == "projects":
        projects = db.query(Project).filter(Project.name.like(search_term)).limit(10).all()
        results.extend([{"type": "project", "id": p.id, "name": p.name, "link": f"/sbase/{p.id}"} for p in projects])

    if not type or type == "workers":
        workers = db.query(Worker).filter(Worker.name.like(search_term)).limit(10).all()
        results.extend([{"type": "worker", "id": w.id, "name": w.name, "link": f"/workers/{w.id}"} for w in workers])

    if not type or type == "vendors":
        vendors = db.query(Vendor).filter(Vendor.name.like(search_term)).limit(10).all()
        results.extend([{"type": "vendor", "id": v.id, "name": v.name, "link": f"/vendors/{v.id}"} for v in vendors])

    if not type or type == "clients":
        clients = db.query(Client).filter(Client.name.like(search_term)).limit(10).all()
        results.extend([{"type": "client", "id": c.id, "name": c.name, "link": f"/clients/{c.id}"} for c in clients])

    return results


# ========== Dashboard Summary API ==========
@app.get("/api/dashboard/summary")
def get_dashboard_summary(db: Session = Depends(get_db)):
    active_projects = db.query(Project).filter(Project.status.in_(["施工中", "受注確定"])).count()
    pending_approvals = db.query(Approval).filter(Approval.status == "pending").count()
    low_stock = db.query(InventoryItem).filter(InventoryItem.quantity <= InventoryItem.min_quantity).count()
    unread_notifications = db.query(Notification).filter(Notification.is_read == False).count()

    return {
        "active_projects": active_projects,
        "pending_approvals": pending_approvals,
        "low_stock": low_stock,
        "unread_notifications": unread_notifications
    }


# ============================================
# タスク17: 図面管理 API
# ============================================

class DrawingCreate(BaseModel):
    project_id: int
    name: str
    file_path: Optional[str] = None
    file_type: Optional[str] = None
    uploaded_by: Optional[str] = None

class DrawingPinCreate(BaseModel):
    drawing_id: int
    x: float
    y: float
    pin_type: Optional[str] = None
    content: Optional[str] = None
    photo_id: Optional[int] = None
    created_by: Optional[str] = None

@app.get("/api/drawings/")
def get_drawings(project_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Drawing)
    if project_id:
        query = query.filter(Drawing.project_id == project_id)
    return query.order_by(Drawing.uploaded_at.desc()).all()

@app.get("/api/drawings/{drawing_id}")
def get_drawing(drawing_id: int, db: Session = Depends(get_db)):
    drawing = db.query(Drawing).filter(Drawing.id == drawing_id).first()
    if not drawing:
        raise HTTPException(status_code=404, detail="Drawing not found")
    return drawing

@app.post("/api/drawings/")
def create_drawing(data: DrawingCreate, db: Session = Depends(get_db)):
    # 同名図面の古いバージョンをis_latest=Falseに
    existing = db.query(Drawing).filter(
        Drawing.project_id == data.project_id,
        Drawing.name == data.name,
        Drawing.is_latest == True
    ).first()
    version = 1
    if existing:
        existing.is_latest = False
        version = existing.version + 1

    drawing = Drawing(**data.model_dump(), version=version)
    db.add(drawing)
    db.commit()
    db.refresh(drawing)
    return drawing

@app.delete("/api/drawings/{drawing_id}")
def delete_drawing(drawing_id: int, db: Session = Depends(get_db)):
    drawing = db.query(Drawing).filter(Drawing.id == drawing_id).first()
    if drawing:
        db.query(DrawingPin).filter(DrawingPin.drawing_id == drawing_id).delete()
        db.delete(drawing)
        db.commit()
    return {"message": "deleted"}

@app.get("/api/drawings/{drawing_id}/pins")
def get_drawing_pins(drawing_id: int, db: Session = Depends(get_db)):
    return db.query(DrawingPin).filter(DrawingPin.drawing_id == drawing_id).all()

@app.post("/api/drawings/pins")
def create_drawing_pin(data: DrawingPinCreate, db: Session = Depends(get_db)):
    pin = DrawingPin(**data.model_dump())
    db.add(pin)
    db.commit()
    db.refresh(pin)
    return pin

@app.delete("/api/drawings/pins/{pin_id}")
def delete_drawing_pin(pin_id: int, db: Session = Depends(get_db)):
    pin = db.query(DrawingPin).filter(DrawingPin.id == pin_id).first()
    if pin:
        db.delete(pin)
        db.commit()
    return {"message": "deleted"}


# ============================================
# タスク18: 工事写真管理 API
# ============================================

class SitePhotoCreate(BaseModel):
    project_id: int
    category: Optional[str] = None
    work_type: Optional[str] = None
    photo_path: Optional[str] = None
    thumbnail_path: Optional[str] = None
    blackboard_data: Optional[str] = None
    taken_at: Optional[datetime] = None
    taken_by: Optional[str] = None
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None
    tags: Optional[str] = None

class BlackboardTemplateCreate(BaseModel):
    name: str
    layout: Optional[str] = None
    fields: Optional[str] = None

@app.get("/api/site-photos/")
def get_site_photos(
    project_id: Optional[int] = None,
    category: Optional[str] = None,
    work_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(SitePhoto)
    if project_id:
        query = query.filter(SitePhoto.project_id == project_id)
    if category:
        query = query.filter(SitePhoto.category == category)
    if work_type:
        query = query.filter(SitePhoto.work_type == work_type)
    return query.order_by(SitePhoto.taken_at.desc()).all()

@app.post("/api/site-photos/")
def create_site_photo(data: SitePhotoCreate, db: Session = Depends(get_db)):
    photo = SitePhoto(**data.model_dump())
    db.add(photo)
    db.commit()
    db.refresh(photo)
    return photo

@app.post("/api/site-photos/bulk")
def create_site_photos_bulk(photos: List[SitePhotoCreate], db: Session = Depends(get_db)):
    created = []
    for p in photos:
        photo = SitePhoto(**p.model_dump())
        db.add(photo)
        created.append(photo)
    db.commit()
    return {"message": f"Created {len(created)} photos"}

@app.delete("/api/site-photos/{photo_id}")
def delete_site_photo(photo_id: int, db: Session = Depends(get_db)):
    photo = db.query(SitePhoto).filter(SitePhoto.id == photo_id).first()
    if photo:
        db.delete(photo)
        db.commit()
    return {"message": "deleted"}

@app.get("/api/blackboard-templates/")
def get_blackboard_templates(db: Session = Depends(get_db)):
    return db.query(BlackboardTemplate).all()

@app.post("/api/blackboard-templates/")
def create_blackboard_template(data: BlackboardTemplateCreate, db: Session = Depends(get_db)):
    template = BlackboardTemplate(**data.model_dump())
    db.add(template)
    db.commit()
    db.refresh(template)
    return template


# ============================================
# タスク19: 検査・是正管理 API
# ============================================

class InspectionCreate(BaseModel):
    project_id: int
    type: Optional[str] = None
    scheduled_date: Optional[date] = None
    completed_date: Optional[date] = None
    inspector: Optional[str] = None
    result: Optional[str] = "pending"
    note: Optional[str] = None

class InspectionItemCreate(BaseModel):
    inspection_id: int
    item_name: str
    standard: Optional[str] = None
    result: Optional[str] = None
    photo_id: Optional[int] = None
    comment: Optional[str] = None

class CorrectionCreate(BaseModel):
    inspection_id: int
    inspection_item_id: Optional[int] = None
    description: Optional[str] = None
    photo_before: Optional[int] = None
    photo_after: Optional[int] = None
    status: Optional[str] = "open"
    due_date: Optional[date] = None
    assigned_to: Optional[str] = None

@app.get("/api/inspections/")
def get_inspections(project_id: Optional[int] = None, result: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Inspection)
    if project_id:
        query = query.filter(Inspection.project_id == project_id)
    if result:
        query = query.filter(Inspection.result == result)
    return query.order_by(Inspection.scheduled_date.desc()).all()

@app.get("/api/inspections/{inspection_id}")
def get_inspection(inspection_id: int, db: Session = Depends(get_db)):
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection not found")
    items = db.query(InspectionItem).filter(InspectionItem.inspection_id == inspection_id).all()
    corrections = db.query(Correction).filter(Correction.inspection_id == inspection_id).all()
    return {"inspection": inspection, "items": items, "corrections": corrections}

@app.post("/api/inspections/")
def create_inspection(data: InspectionCreate, db: Session = Depends(get_db)):
    inspection = Inspection(**data.model_dump())
    db.add(inspection)
    db.commit()
    db.refresh(inspection)
    return inspection

@app.put("/api/inspections/{inspection_id}")
def update_inspection(inspection_id: int, data: InspectionCreate, db: Session = Depends(get_db)):
    inspection = db.query(Inspection).filter(Inspection.id == inspection_id).first()
    if not inspection:
        raise HTTPException(status_code=404, detail="Inspection not found")
    for key, value in data.model_dump().items():
        setattr(inspection, key, value)
    db.commit()
    return inspection

@app.post("/api/inspection-items/")
def create_inspection_item(data: InspectionItemCreate, db: Session = Depends(get_db)):
    item = InspectionItem(**data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item

@app.put("/api/inspection-items/{item_id}")
def update_inspection_item(item_id: int, data: dict, db: Session = Depends(get_db)):
    item = db.query(InspectionItem).filter(InspectionItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Inspection item not found")
    for key, value in data.items():
        setattr(item, key, value)
    db.commit()
    return item

@app.get("/api/corrections/")
def get_corrections(status: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Correction)
    if status:
        query = query.filter(Correction.status == status)
    return query.order_by(Correction.due_date).all()

@app.post("/api/corrections/")
def create_correction(data: CorrectionCreate, db: Session = Depends(get_db)):
    correction = Correction(**data.model_dump())
    db.add(correction)
    db.commit()
    db.refresh(correction)
    return correction

@app.put("/api/corrections/{correction_id}")
def update_correction(correction_id: int, data: dict, db: Session = Depends(get_db)):
    correction = db.query(Correction).filter(Correction.id == correction_id).first()
    if not correction:
        raise HTTPException(status_code=404, detail="Correction not found")
    for key, value in data.items():
        if key in ['due_date', 'completed_date'] and value:
            value = datetime.strptime(value, '%Y-%m-%d').date() if isinstance(value, str) else value
        setattr(correction, key, value)
    db.commit()
    return correction


# ============================================
# タスク20: 安全書類（グリーンファイル）API
# ============================================

class WorkerRegistrationCreate(BaseModel):
    worker_id: int
    project_id: int
    blood_type: Optional[str] = None
    emergency_contact: Optional[str] = None
    emergency_phone: Optional[str] = None
    qualifications: Optional[str] = None
    health_check_date: Optional[date] = None
    insurance_number: Optional[str] = None

class SafetyTrainingCreate(BaseModel):
    project_id: int
    worker_id: int
    training_date: Optional[date] = None
    training_type: Optional[str] = None
    trainer: Optional[str] = None
    signature_path: Optional[str] = None

class QualificationCreate(BaseModel):
    worker_id: int
    name: str
    number: Optional[str] = None
    issue_date: Optional[date] = None
    expiry_date: Optional[date] = None
    photo_path: Optional[str] = None

@app.get("/api/worker-registrations/")
def get_worker_registrations(project_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(WorkerRegistration)
    if project_id:
        query = query.filter(WorkerRegistration.project_id == project_id)
    registrations = query.all()
    result = []
    for reg in registrations:
        worker = db.query(Worker).filter(Worker.id == reg.worker_id).first()
        result.append({
            **reg.__dict__,
            "worker_name": worker.name if worker else None
        })
    return result

@app.post("/api/worker-registrations/")
def create_worker_registration(data: WorkerRegistrationCreate, db: Session = Depends(get_db)):
    reg = WorkerRegistration(**data.model_dump())
    db.add(reg)
    db.commit()
    db.refresh(reg)
    return reg

@app.put("/api/worker-registrations/{reg_id}")
def update_worker_registration(reg_id: int, data: WorkerRegistrationCreate, db: Session = Depends(get_db)):
    reg = db.query(WorkerRegistration).filter(WorkerRegistration.id == reg_id).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Registration not found")
    for key, value in data.model_dump().items():
        setattr(reg, key, value)
    db.commit()
    return reg

@app.get("/api/safety-trainings/")
def get_safety_trainings(project_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(SafetyTraining)
    if project_id:
        query = query.filter(SafetyTraining.project_id == project_id)
    return query.order_by(SafetyTraining.training_date.desc()).all()

@app.post("/api/safety-trainings/")
def create_safety_training(data: SafetyTrainingCreate, db: Session = Depends(get_db)):
    training = SafetyTraining(**data.model_dump())
    db.add(training)
    db.commit()
    db.refresh(training)
    return training

@app.get("/api/qualifications/")
def get_qualifications(worker_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Qualification)
    if worker_id:
        query = query.filter(Qualification.worker_id == worker_id)
    return query.all()

@app.get("/api/qualifications/expiring")
def get_expiring_qualifications(db: Session = Depends(get_db)):
    alert_date = date.today() + relativedelta(days=30)
    return db.query(Qualification).filter(Qualification.expiry_date <= alert_date).all()

@app.post("/api/qualifications/")
def create_qualification(data: QualificationCreate, db: Session = Depends(get_db)):
    qual = Qualification(**data.model_dump())
    db.add(qual)
    db.commit()
    db.refresh(qual)
    return qual

@app.get("/api/workers/{worker_id}/safety-documents")
def get_worker_safety_documents(worker_id: int, db: Session = Depends(get_db)):
    registrations = db.query(WorkerRegistration).filter(WorkerRegistration.worker_id == worker_id).all()
    trainings = db.query(SafetyTraining).filter(SafetyTraining.worker_id == worker_id).all()
    qualifications = db.query(Qualification).filter(Qualification.worker_id == worker_id).all()
    return {"registrations": registrations, "trainings": trainings, "qualifications": qualifications}

@app.get("/api/safety/health-check-alerts")
def get_health_check_alerts(db: Session = Depends(get_db)):
    """健康診断期限が近い作業員を取得"""
    alert_date = date.today() + relativedelta(months=1)
    registrations = db.query(WorkerRegistration).filter(
        WorkerRegistration.health_check_date <= alert_date
    ).all()
    result = []
    for reg in registrations:
        worker = db.query(Worker).filter(Worker.id == reg.worker_id).first()
        result.append({
            "worker_id": reg.worker_id,
            "worker_name": worker.name if worker else None,
            "health_check_date": reg.health_check_date.isoformat() if reg.health_check_date else None
        })
    return result


# ============================================
# タスク21: 案件別チャット API
# ============================================

class MessageCreate(BaseModel):
    project_id: int
    sender_id: Optional[str] = None
    sender_name: Optional[str] = None
    content: Optional[str] = None
    attachment_path: Optional[str] = None
    attachment_type: Optional[str] = None

@app.get("/api/messages/")
def get_messages(project_id: int, limit: int = 50, offset: int = 0, db: Session = Depends(get_db)):
    messages = db.query(Message).filter(Message.project_id == project_id).order_by(Message.sent_at.desc()).offset(offset).limit(limit).all()
    return messages[::-1]  # 古い順に返す

@app.post("/api/messages/")
def create_message(data: MessageCreate, db: Session = Depends(get_db)):
    message = Message(**data.model_dump())
    db.add(message)
    db.commit()
    db.refresh(message)
    return message

@app.get("/api/messages/unread-count")
def get_unread_message_count(project_id: Optional[int] = None, user_id: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Message).filter(Message.is_read == False)
    if project_id:
        query = query.filter(Message.project_id == project_id)
    return {"count": query.count()}

@app.put("/api/messages/{message_id}/read")
def mark_message_read(message_id: int, user_id: str, db: Session = Depends(get_db)):
    read = MessageRead(message_id=message_id, user_id=user_id)
    db.add(read)
    message = db.query(Message).filter(Message.id == message_id).first()
    if message:
        message.is_read = True
    db.commit()
    return {"message": "marked as read"}

@app.put("/api/messages/read-all")
def mark_all_messages_read(project_id: int, user_id: str, db: Session = Depends(get_db)):
    db.query(Message).filter(Message.project_id == project_id, Message.is_read == False).update({"is_read": True})
    db.commit()
    return {"message": "all marked as read"}


# ============================================
# タスク22: 権限管理 API
# ============================================

class UserCreate(BaseModel):
    email: str
    name: str
    role: Optional[str] = "worker"
    department: Optional[str] = None
    is_active: Optional[bool] = True

class PermissionCreate(BaseModel):
    role: str
    resource: str
    action: str
    is_allowed: Optional[bool] = True

@app.get("/api/users/")
def get_users(role: Optional[str] = None, is_active: Optional[bool] = True, db: Session = Depends(get_db)):
    query = db.query(User)
    if role:
        query = query.filter(User.role == role)
    if is_active is not None:
        query = query.filter(User.is_active == is_active)
    return query.order_by(User.name).all()

@app.get("/api/users/{user_id}")
def get_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@app.post("/api/users/")
def create_user(data: UserCreate, db: Session = Depends(get_db)):
    user = User(**data.model_dump())
    db.add(user)
    db.commit()
    db.refresh(user)
    return user

@app.put("/api/users/{user_id}")
def update_user(user_id: int, data: UserCreate, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    for key, value in data.model_dump().items():
        setattr(user, key, value)
    db.commit()
    return user

@app.delete("/api/users/{user_id}")
def delete_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if user:
        user.is_active = False
        db.commit()
    return {"message": "deactivated"}

@app.get("/api/permissions/")
def get_permissions(role: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Permission)
    if role:
        query = query.filter(Permission.role == role)
    return query.all()

@app.post("/api/permissions/")
def create_permission(data: PermissionCreate, db: Session = Depends(get_db)):
    # 既存があれば更新
    existing = db.query(Permission).filter(
        Permission.role == data.role,
        Permission.resource == data.resource,
        Permission.action == data.action
    ).first()
    if existing:
        existing.is_allowed = data.is_allowed
    else:
        perm = Permission(**data.model_dump())
        db.add(perm)
    db.commit()
    return {"message": "updated"}

@app.get("/api/permissions/check")
def check_permission(role: str, resource: str, action: str, db: Session = Depends(get_db)):
    # adminは全権限
    if role == "admin":
        return {"allowed": True}
    perm = db.query(Permission).filter(
        Permission.role == role,
        Permission.resource == resource,
        Permission.action == action
    ).first()
    return {"allowed": perm.is_allowed if perm else False}


# ============================================
# タスク23: 外部連携 API
# ============================================

import csv
from io import StringIO
from fastapi.responses import StreamingResponse

class IntegrationSettingCreate(BaseModel):
    service: str
    api_key: Optional[str] = None
    access_token: Optional[str] = None
    config: Optional[str] = None
    is_active: Optional[bool] = False

# CSV出力（弥生会計フォーマット）
@app.get("/api/export/yayoi")
def export_yayoi(year_month: str, db: Session = Depends(get_db)):
    year, month = map(int, year_month.split('-'))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    costs = db.query(Cost).filter(Cost.date >= start_date, Cost.date < end_date).all()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["日付", "借方科目", "借方金額", "貸方科目", "貸方金額", "摘要"])

    category_accounts = {
        "労務費": "労務費",
        "材料費": "材料費",
        "外注費": "外注費",
        "経費": "諸経費"
    }

    for cost in costs:
        account = category_accounts.get(cost.category, "諸経費")
        writer.writerow([
            cost.date.strftime("%Y/%m/%d") if cost.date else "",
            account,
            cost.amount or 0,
            "未払金",
            cost.amount or 0,
            f"{cost.vendor or ''} {cost.description or ''}"
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=yayoi_{year_month}.csv"}
    )

# CSV出力（freeeフォーマット）
@app.get("/api/export/freee")
def export_freee(year_month: str, db: Session = Depends(get_db)):
    year, month = map(int, year_month.split('-'))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    costs = db.query(Cost).filter(Cost.date >= start_date, Cost.date < end_date).all()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["取引日", "勘定科目", "税区分", "金額", "取引先", "品目", "メモ"])

    for cost in costs:
        writer.writerow([
            cost.date.strftime("%Y-%m-%d") if cost.date else "",
            cost.category or "",
            "課税仕入10%",
            cost.amount or 0,
            cost.vendor or "",
            cost.work_type or "",
            cost.description or ""
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=freee_{year_month}.csv"}
    )

# CSV出力（マネーフォワードフォーマット）
@app.get("/api/export/moneyforward")
def export_moneyforward(year_month: str, db: Session = Depends(get_db)):
    year, month = map(int, year_month.split('-'))
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)

    costs = db.query(Cost).filter(Cost.date >= start_date, Cost.date < end_date).all()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["日付", "借方勘定科目", "借方補助科目", "借方金額", "貸方勘定科目", "貸方補助科目", "貸方金額", "摘要"])

    for cost in costs:
        writer.writerow([
            cost.date.strftime("%Y/%m/%d") if cost.date else "",
            cost.category or "",
            cost.work_type or "",
            cost.amount or 0,
            "未払金",
            cost.vendor or "",
            cost.amount or 0,
            cost.description or ""
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=moneyforward_{year_month}.csv"}
    )

# 外部連携設定
@app.get("/api/integrations/")
def get_integrations(db: Session = Depends(get_db)):
    return db.query(IntegrationSetting).all()

@app.post("/api/integrations/")
def create_integration(data: IntegrationSettingCreate, db: Session = Depends(get_db)):
    existing = db.query(IntegrationSetting).filter(IntegrationSetting.service == data.service).first()
    if existing:
        for key, value in data.model_dump().items():
            setattr(existing, key, value)
        db.commit()
        return existing
    setting = IntegrationSetting(**data.model_dump())
    db.add(setting)
    db.commit()
    db.refresh(setting)
    return setting

# LINE通知（設定があれば送信）
@app.post("/api/notify/line")
async def send_line_notification(message: str, db: Session = Depends(get_db)):
    setting = db.query(IntegrationSetting).filter(
        IntegrationSetting.service == "line",
        IntegrationSetting.is_active == True
    ).first()
    if not setting or not setting.access_token:
        return {"success": False, "error": "LINE連携が設定されていません"}

    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://notify-api.line.me/api/notify",
                headers={"Authorization": f"Bearer {setting.access_token}"},
                data={"message": message}
            )
            return {"success": response.status_code == 200}
    except Exception as e:
        return {"success": False, "error": str(e)}

# Googleカレンダー連携用エンドポイント
@app.get("/api/calendar/events")
def get_calendar_events(year_month: str, db: Session = Depends(get_db)):
    """工程をカレンダーイベント形式で取得"""
    schedules = db.query(Schedule).all()
    events = []
    for s in schedules:
        project = db.query(Project).filter(Project.id == s.project_id).first()
        events.append({
            "id": s.id,
            "title": project.name if project else f"Project {s.project_id}",
            "start": s.start_date.isoformat() if s.start_date else None,
            "end": s.end_date.isoformat() if s.end_date else None,
            "color": s.color,
            "project_id": s.project_id
        })
    return events


# ============================================
# タスク7-2: 日報 (Daily Reports) API
# ============================================

class DailyReportCreate(BaseModel):
    date: date
    worker_id: int
    project_id: Optional[int] = None
    hours: Optional[float] = 8
    overtime_hours: Optional[float] = 0
    note: Optional[str] = None

@app.get("/api/daily-reports/")
def get_daily_reports(
    project_id: Optional[int] = None,
    worker_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(DailyReport)
    if project_id:
        query = query.filter(DailyReport.project_id == project_id)
    if worker_id:
        query = query.filter(DailyReport.worker_id == worker_id)
    if date_from:
        query = query.filter(DailyReport.date >= date_from)
    if date_to:
        query = query.filter(DailyReport.date <= date_to)
    return query.order_by(DailyReport.date.desc()).all()

@app.post("/api/daily-reports/")
def create_daily_report(data: DailyReportCreate, db: Session = Depends(get_db)):
    report = DailyReport(**data.model_dump())
    db.add(report)
    db.commit()
    db.refresh(report)
    return report

@app.put("/api/daily-reports/{report_id}")
def update_daily_report(report_id: int, data: DailyReportCreate, db: Session = Depends(get_db)):
    report = db.query(DailyReport).filter(DailyReport.id == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    for key, value in data.model_dump().items():
        setattr(report, key, value)
    db.commit()
    return report

@app.delete("/api/daily-reports/{report_id}")
def delete_daily_report(report_id: int, db: Session = Depends(get_db)):
    report = db.query(DailyReport).filter(DailyReport.id == report_id).first()
    if report:
        db.delete(report)
        db.commit()
    return {"ok": True}

@app.post("/api/daily-reports/generate-from-assignments")
def generate_reports_from_assignments(target_date: date, db: Session = Depends(get_db)):
    """段取りくんの配置から日報を自動生成"""
    assignments = db.query(Assignment).filter(Assignment.date == target_date).all()
    created = 0
    for a in assignments:
        existing = db.query(DailyReport).filter(
            DailyReport.date == target_date,
            DailyReport.worker_id == a.worker_id,
            DailyReport.project_id == a.project_id
        ).first()
        if not existing:
            report = DailyReport(
                date=target_date,
                worker_id=a.worker_id,
                project_id=a.project_id,
                hours=8,
                overtime_hours=0,
                note=a.note
            )
            db.add(report)
            created += 1
    db.commit()
    return {"created": created}

@app.get("/api/projects/{project_id}/labor-cost")
def get_project_labor_cost(project_id: int, db: Session = Depends(get_db)):
    """案件の労務費を日報から自動計算"""
    reports = db.query(DailyReport).filter(DailyReport.project_id == project_id).all()
    total_labor_cost = 0
    for r in reports:
        worker = db.query(Worker).filter(Worker.id == r.worker_id).first()
        if worker and worker.daily_rate:
            daily_cost = worker.daily_rate * (r.hours / 8)
            overtime_cost = (worker.daily_rate / 8) * 1.25 * (r.overtime_hours or 0)
            total_labor_cost += daily_cost + overtime_cost
    return {
        "project_id": project_id,
        "labor_cost": int(total_labor_cost),
        "report_count": len(reports)
    }


# ============================================
# タスク7-7: 会社設定 (Company Settings) API
# ============================================

class CompanySettingsUpdate(BaseModel):
    company_name: Optional[str] = None
    postal_code: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    fax: Optional[str] = None
    email: Optional[str] = None
    invoice_number: Optional[str] = None
    logo_path: Optional[str] = None
    bank_name: Optional[str] = None
    bank_branch: Optional[str] = None
    account_type: Optional[str] = None
    account_number: Optional[str] = None
    account_name: Optional[str] = None
    fiscal_year_start: Optional[int] = 4
    annual_target: Optional[float] = 0

@app.get("/api/company-settings/")
def get_company_settings(db: Session = Depends(get_db)):
    settings = db.query(CompanySettings).first()
    if not settings:
        settings = CompanySettings(company_name="サンユウテック")
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return settings

@app.put("/api/company-settings/")
def update_company_settings(data: CompanySettingsUpdate, db: Session = Depends(get_db)):
    settings = db.query(CompanySettings).first()
    if not settings:
        settings = CompanySettings()
        db.add(settings)
    for key, value in data.model_dump().items():
        if value is not None:
            setattr(settings, key, value)
    db.commit()
    db.refresh(settings)
    return settings


# ============================================
# タスク7-6: 帳票電子発行 (Document Send) API
# ============================================

class DocumentSendCreate(BaseModel):
    document_type: str
    document_id: int
    recipient_email: str

@app.get("/api/document-sends/")
def get_document_sends(document_type: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(DocumentSend)
    if document_type:
        query = query.filter(DocumentSend.document_type == document_type)
    return query.order_by(DocumentSend.sent_at.desc()).all()

@app.post("/api/document-sends/")
def create_document_send(data: DocumentSendCreate, db: Session = Depends(get_db)):
    send = DocumentSend(**data.model_dump())
    db.add(send)
    db.commit()
    db.refresh(send)
    return send

@app.get("/api/document-sends/{send_id}/track")
def track_document_open(send_id: int, db: Session = Depends(get_db)):
    """メール開封トラッキング"""
    send = db.query(DocumentSend).filter(DocumentSend.id == send_id).first()
    if send and not send.opened_at:
        send.opened_at = datetime.now()
        send.status = "opened"
        db.commit()
    return {"ok": True}


# ============================================
# タスク7-5: PDF生成 API
# ============================================

from fastapi.responses import StreamingResponse
from io import BytesIO

@app.get("/api/documents/estimate/{estimate_id}/pdf")
def generate_estimate_pdf(estimate_id: int, db: Session = Depends(get_db)):
    """見積書PDF生成（簡易版 - HTML形式で返す）"""
    # 実際のPDF生成はfpdf2やreportlabなどを使用
    # ここでは簡易的にJSONで必要データを返す
    estimate = db.query(Estimate).filter(Estimate.id == estimate_id).first()
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")
    project = db.query(Project).filter(Project.id == estimate.project_id).first()
    company = db.query(CompanySettings).first()
    return {
        "type": "estimate",
        "project": {"id": project.id, "name": project.name, "client": project.client} if project else None,
        "estimate": {
            "work_type": estimate.work_type,
            "description": estimate.description,
            "quantity": estimate.quantity,
            "unit": estimate.unit,
            "unit_price": estimate.unit_price,
            "amount": estimate.amount
        },
        "company": {
            "name": company.company_name if company else "サンユウテック",
            "address": company.address if company else "",
            "phone": company.phone if company else "",
            "invoice_number": company.invoice_number if company else ""
        }
    }

@app.get("/api/documents/project/{project_id}/ledger")
def get_project_ledger(project_id: int, db: Session = Depends(get_db)):
    """工事台帳データ取得（タスク7-1対応）"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # 原価集計
    costs = db.query(Cost).filter(Cost.project_id == project_id).all()
    cost_by_category = {"労務費": 0, "材料費": 0, "外注費": 0, "機械費": 0, "経費": 0}
    for c in costs:
        if c.category in cost_by_category:
            cost_by_category[c.category] += c.amount or 0

    # 日報からの労務費
    reports = db.query(DailyReport).filter(DailyReport.project_id == project_id).all()
    labor_from_reports = 0
    for r in reports:
        worker = db.query(Worker).filter(Worker.id == r.worker_id).first()
        if worker and worker.daily_rate:
            labor_from_reports += worker.daily_rate * (r.hours / 8)
            labor_from_reports += (worker.daily_rate / 8) * 1.25 * (r.overtime_hours or 0)

    # 出来高累計
    progress = db.query(MonthlyProgress).filter(MonthlyProgress.project_id == project_id).all()
    total_progress = sum(p.progress_amount or 0 for p in progress)

    # 請求金額累計
    billings = db.query(Billing).filter(Billing.project_id == project_id).all()
    total_billed = sum(b.amount or 0 for b in billings)

    total_cost = sum(cost_by_category.values())
    order_amount = project.order_amount or 0

    return {
        "project": {
            "id": project.id,
            "code": project.code,
            "name": project.name,
            "client": project.client,
            "order_amount": order_amount,
            "order_amount_with_tax": int(order_amount * (1 + (project.tax_rate or 0.1))),
            "budget_amount": project.budget_amount
        },
        "costs": {
            **cost_by_category,
            "labor_from_reports": int(labor_from_reports),
            "total": total_cost
        },
        "billing": {
            "total_billed": total_billed,
            "total_progress": total_progress
        },
        "profit": {
            "gross_profit": total_billed - total_cost,
            "gross_profit_rate": round((total_billed - total_cost) / total_billed * 100, 1) if total_billed > 0 else 0,
            "progress_rate": round(total_progress / order_amount * 100, 1) if order_amount > 0 else 0
        }
    }


# ============================================
# タスク7-3: 経営分析 (Analytics) API
# ============================================

@app.get("/api/analytics/monthly-sales")
def get_monthly_sales(year: Optional[int] = None, db: Session = Depends(get_db)):
    """月別売上推移"""
    if not year:
        year = datetime.now().year

    result = []
    for month in range(1, 13):
        year_month = f"{year}-{month:02d}"
        # 出来高から売上を計算
        progress = db.query(func.sum(MonthlyProgress.progress_amount)).filter(
            MonthlyProgress.year_month == year_month
        ).scalar() or 0

        # 原価を計算
        start_date = date(year, month, 1)
        end_date = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
        costs = db.query(func.sum(Cost.amount)).filter(
            Cost.date >= start_date,
            Cost.date < end_date
        ).scalar() or 0

        result.append({
            "month": year_month,
            "sales": progress,
            "cost": costs,
            "profit": progress - costs
        })

    return result

@app.get("/api/analytics/client-breakdown")
def get_client_breakdown(year: Optional[int] = None, db: Session = Depends(get_db)):
    """顧客別売上比率"""
    if not year:
        year = datetime.now().year

    # プロジェクト別に出来高を集計
    projects = db.query(Project).all()
    client_sales = {}

    for p in projects:
        progress = db.query(func.sum(MonthlyProgress.progress_amount)).filter(
            MonthlyProgress.project_id == p.id,
            MonthlyProgress.year_month.like(f"{year}-%")
        ).scalar() or 0

        client = p.client or "その他"
        if client not in client_sales:
            client_sales[client] = 0
        client_sales[client] += progress

    return [{"client": k, "sales": v} for k, v in sorted(client_sales.items(), key=lambda x: -x[1])]

@app.get("/api/analytics/person-ranking")
def get_person_ranking(year: Optional[int] = None, db: Session = Depends(get_db)):
    """担当者別粗利ランキング"""
    if not year:
        year = datetime.now().year

    projects = db.query(Project).all()
    person_profit = {}

    for p in projects:
        # 出来高
        progress = db.query(func.sum(MonthlyProgress.progress_amount)).filter(
            MonthlyProgress.project_id == p.id,
            MonthlyProgress.year_month.like(f"{year}-%")
        ).scalar() or 0

        # 原価
        costs = db.query(func.sum(Cost.amount)).filter(
            Cost.project_id == p.id
        ).scalar() or 0

        person = p.site_person or p.sales_person or "未割当"
        if person not in person_profit:
            person_profit[person] = {"sales": 0, "cost": 0, "profit": 0}
        person_profit[person]["sales"] += progress
        person_profit[person]["cost"] += costs
        person_profit[person]["profit"] += progress - costs

    return [{"person": k, **v} for k, v in sorted(person_profit.items(), key=lambda x: -x[1]["profit"])]

@app.get("/api/analytics/target-vs-actual")
def get_target_vs_actual(db: Session = Depends(get_db)):
    """年間売上目標vs実績"""
    company = db.query(CompanySettings).first()
    target = company.annual_target if company else 0

    # 今期の売上を計算
    fiscal_start = company.fiscal_year_start if company else 4
    now = datetime.now()
    if now.month >= fiscal_start:
        fiscal_year_start = date(now.year, fiscal_start, 1)
    else:
        fiscal_year_start = date(now.year - 1, fiscal_start, 1)

    actual = db.query(func.sum(MonthlyProgress.progress_amount)).filter(
        MonthlyProgress.year_month >= fiscal_year_start.strftime("%Y-%m")
    ).scalar() or 0

    return {
        "target": target,
        "actual": actual,
        "achievement_rate": round(actual / target * 100, 1) if target > 0 else 0,
        "remaining": target - actual
    }

@app.get("/api/analytics/profit-trend")
def get_profit_trend(year: Optional[int] = None, db: Session = Depends(get_db)):
    """月別粗利率推移"""
    if not year:
        year = datetime.now().year

    result = []
    for month in range(1, 13):
        year_month = f"{year}-{month:02d}"
        progress = db.query(func.sum(MonthlyProgress.progress_amount)).filter(
            MonthlyProgress.year_month == year_month
        ).scalar() or 0

        start_date = date(year, month, 1)
        end_date = date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
        costs = db.query(func.sum(Cost.amount)).filter(
            Cost.date >= start_date,
            Cost.date < end_date
        ).scalar() or 0

        profit = progress - costs
        profit_rate = round(profit / progress * 100, 1) if progress > 0 else 0

        result.append({
            "month": year_month,
            "profit": profit,
            "profit_rate": profit_rate
        })

    return result


# ============================================
# タスク12: LINE WORKS連携 API
# ============================================

class LineWorksSettingsUpdate(BaseModel):
    bot_id: Optional[str] = None
    client_id: Optional[str] = None
    client_secret: Optional[str] = None
    service_account: Optional[str] = None
    private_key: Optional[str] = None
    is_active: Optional[bool] = True

@app.get("/api/lineworks/settings")
def get_lineworks_settings(db: Session = Depends(get_db)):
    settings = db.query(LineWorksSettings).first()
    if not settings:
        settings = LineWorksSettings()
        db.add(settings)
        db.commit()
        db.refresh(settings)
    return {
        "id": settings.id,
        "bot_id": settings.bot_id,
        "client_id": settings.client_id,
        "client_secret": "***" if settings.client_secret else None,
        "service_account": settings.service_account,
        "has_private_key": bool(settings.private_key),
        "is_active": settings.is_active
    }

@app.put("/api/lineworks/settings")
def update_lineworks_settings(data: LineWorksSettingsUpdate, db: Session = Depends(get_db)):
    settings = db.query(LineWorksSettings).first()
    if not settings:
        settings = LineWorksSettings()
        db.add(settings)
    for key, value in data.model_dump().items():
        if value is not None:
            setattr(settings, key, value)
    db.commit()
    return {"ok": True}

@app.post("/api/lineworks/test-connection")
async def test_lineworks_connection(db: Session = Depends(get_db)):
    """LINE WORKS接続テスト"""
    settings = db.query(LineWorksSettings).first()
    if not settings or not settings.client_id:
        return {"success": False, "error": "LINE WORKS設定が未完了です"}

    # 実際の接続テストはJWTトークン生成が必要
    # ここでは設定の存在確認のみ
    return {
        "success": True,
        "message": "設定が確認できました（実際の接続にはBot認証が必要です）"
    }

@app.get("/api/lineworks/notifications")
def get_lineworks_notifications(db: Session = Depends(get_db)):
    notifications = db.query(LineWorksNotification).all()
    if not notifications:
        # デフォルト通知設定を作成
        defaults = [
            {"type": "approval", "is_enabled": True, "template": "承認依頼があります: {title}"},
            {"type": "daily_reminder", "is_enabled": True, "schedule": "0 18 * * *", "template": "本日の日報を入力してください"},
            {"type": "assignment", "is_enabled": True, "template": "明日の配置: {project} - {time}"},
            {"type": "ky_reminder", "is_enabled": True, "schedule": "0 8 * * *", "template": "KY活動を実施してください"},
            {"type": "inventory_alert", "is_enabled": True, "template": "在庫少: {item} 残り{quantity}"},
            {"type": "weather_alert", "is_enabled": True, "template": "{project}: {weather}予報 降水確率{probability}%"},
        ]
        for d in defaults:
            n = LineWorksNotification(**d)
            db.add(n)
        db.commit()
        notifications = db.query(LineWorksNotification).all()
    return notifications

@app.put("/api/lineworks/notifications/{notification_id}")
def update_lineworks_notification(notification_id: int, data: dict, db: Session = Depends(get_db)):
    notification = db.query(LineWorksNotification).filter(LineWorksNotification.id == notification_id).first()
    if notification:
        for key, value in data.items():
            setattr(notification, key, value)
        db.commit()
    return {"ok": True}

@app.get("/api/lineworks/users")
def get_lineworks_users(db: Session = Depends(get_db)):
    users = db.query(LineWorksUser).all()
    result = []
    for u in users:
        user = db.query(User).filter(User.id == u.user_id).first()
        result.append({
            "id": u.id,
            "user_id": u.user_id,
            "user_name": user.name if user else None,
            "lineworks_user_id": u.lineworks_user_id,
            "is_active": u.is_active
        })
    return result

@app.post("/api/lineworks/users")
def create_lineworks_user(data: dict, db: Session = Depends(get_db)):
    user = LineWorksUser(
        user_id=data.get("user_id"),
        lineworks_user_id=data.get("lineworks_user_id"),
        is_active=True
    )
    db.add(user)
    db.commit()
    return {"ok": True}

@app.get("/api/lineworks/logs")
def get_lineworks_logs(limit: int = 50, db: Session = Depends(get_db)):
    return db.query(LineWorksLog).order_by(LineWorksLog.sent_at.desc()).limit(limit).all()

@app.post("/api/lineworks/send")
async def send_lineworks_message(data: dict, db: Session = Depends(get_db)):
    """LINE WORKSメッセージ送信"""
    settings = db.query(LineWorksSettings).first()
    if not settings or not settings.is_active:
        return {"success": False, "error": "LINE WORKSが無効です"}

    # 送信ログを記録
    log = LineWorksLog(
        type=data.get("type", "manual"),
        recipient=data.get("recipient"),
        message=data.get("message"),
        status="sent"
    )
    db.add(log)
    db.commit()

    # 実際のメッセージ送信はLINE WORKS APIを呼び出す
    # ここでは簡易実装
    return {"success": True, "log_id": log.id}

@app.post("/api/lineworks/webhook")
async def lineworks_webhook(data: dict, db: Session = Depends(get_db)):
    """LINE WORKSからのWebhook受信（双方向通信）"""
    message = data.get("content", {}).get("text", "")
    user_id = data.get("source", {}).get("userId", "")

    response_message = ""

    if "今日の配置" in message:
        today = date.today()
        assignments = db.query(Assignment).filter(Assignment.date == today).all()
        if assignments:
            response_message = "【本日の配置】\n"
            for a in assignments:
                worker = db.query(Worker).filter(Worker.id == a.worker_id).first()
                project = db.query(Project).filter(Project.id == a.project_id).first()
                response_message += f"・{worker.name if worker else '?'}: {project.name if project else '?'}\n"
        else:
            response_message = "本日の配置情報はありません"

    elif "明日の配置" in message:
        tomorrow = date.today() + relativedelta(days=1)
        assignments = db.query(Assignment).filter(Assignment.date == tomorrow).all()
        if assignments:
            response_message = "【明日の配置】\n"
            for a in assignments:
                worker = db.query(Worker).filter(Worker.id == a.worker_id).first()
                project = db.query(Project).filter(Project.id == a.project_id).first()
                response_message += f"・{worker.name if worker else '?'}: {project.name if project else '?'}\n"
        else:
            response_message = "明日の配置情報はありません"

    elif "承認一覧" in message:
        approvals = db.query(Approval).filter(Approval.status == "pending").all()
        if approvals:
            response_message = f"【承認待ち: {len(approvals)}件】\n"
            for a in approvals[:5]:
                response_message += f"・{a.type} ID:{a.reference_id}\n"
        else:
            response_message = "承認待ちはありません"

    elif message.startswith("承認 "):
        try:
            approval_id = int(message.replace("承認 ", "").strip())
            approval = db.query(Approval).filter(Approval.id == approval_id).first()
            if approval:
                approval.status = "approved"
                approval.approved_at = datetime.now()
                db.commit()
                response_message = f"ID:{approval_id}を承認しました"
            else:
                response_message = "該当する承認が見つかりません"
        except:
            response_message = "承認コマンドの形式: 承認 {ID}"

    elif "天気" in message:
        projects = db.query(Project).filter(Project.status == "施工中").limit(3).all()
        response_message = "【現場天気】\n天気情報はアプリで確認してください"

    elif "日報" in message:
        response_message = "日報入力はこちら: https://sbase.sanyutech.jp/daily-reports"

    else:
        response_message = "コマンド一覧:\n・今日の配置\n・明日の配置\n・承認一覧\n・承認 {ID}\n・天気\n・日報"

    return {"message": response_message}


# ============================================
# 名刺図書館 API
# ============================================

@app.get("/api/business-cards/")
def get_business_cards(
    search: Optional[str] = None,
    tag: Optional[str] = None,
    favorite_only: bool = False,
    db: Session = Depends(get_db)
):
    """名刺一覧取得"""
    query = db.query(BusinessCard).order_by(BusinessCard.company_name, BusinessCard.person_name)

    if search:
        query = query.filter(
            (BusinessCard.company_name.contains(search)) |
            (BusinessCard.person_name.contains(search)) |
            (BusinessCard.memo.contains(search))
        )

    if tag:
        query = query.filter(BusinessCard.tag == tag)

    if favorite_only:
        query = query.filter(BusinessCard.is_favorite == True)

    cards = query.all()

    # 会社別にグループ化
    grouped = {}
    for card in cards:
        company = card.company_name or "その他"
        if company not in grouped:
            grouped[company] = []
        grouped[company].append({
            "id": card.id,
            "company_name": card.company_name,
            "person_name": card.person_name,
            "department": card.department,
            "position": card.position,
            "phone": card.phone,
            "mobile": card.mobile,
            "email": card.email,
            "address": card.address,
            "url": card.url,
            "image_path": card.image_path,
            "tag": card.tag,
            "is_favorite": card.is_favorite,
            "memo": card.memo,
            "project_ids": card.project_ids,
            "created_at": card.created_at
        })

    return {"cards": cards, "grouped": grouped, "total": len(cards)}


@app.get("/api/business-cards/{card_id}")
def get_business_card(card_id: int, db: Session = Depends(get_db)):
    """名刺詳細取得"""
    card = db.query(BusinessCard).filter(BusinessCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")
    return card


@app.post("/api/business-cards/")
def create_business_card(data: dict, db: Session = Depends(get_db)):
    """名刺登録"""
    card = BusinessCard(
        company_name=data.get("company_name"),
        person_name=data.get("person_name"),
        department=data.get("department"),
        position=data.get("position"),
        phone=data.get("phone"),
        mobile=data.get("mobile"),
        email=data.get("email"),
        address=data.get("address"),
        url=data.get("url"),
        image_path=data.get("image_path"),
        tag=data.get("tag", "other"),
        is_favorite=data.get("is_favorite", False),
        memo=data.get("memo"),
        project_ids=data.get("project_ids")
    )
    db.add(card)
    db.commit()
    db.refresh(card)
    return card


@app.put("/api/business-cards/{card_id}")
def update_business_card(card_id: int, data: dict, db: Session = Depends(get_db)):
    """名刺更新"""
    card = db.query(BusinessCard).filter(BusinessCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    for key, value in data.items():
        if hasattr(card, key):
            setattr(card, key, value)

    db.commit()
    db.refresh(card)
    return card


@app.put("/api/business-cards/{card_id}/favorite")
def toggle_favorite(card_id: int, db: Session = Depends(get_db)):
    """お気に入り切替"""
    card = db.query(BusinessCard).filter(BusinessCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    card.is_favorite = not card.is_favorite
    db.commit()
    return {"id": card_id, "is_favorite": card.is_favorite}


@app.delete("/api/business-cards/{card_id}")
def delete_business_card(card_id: int, db: Session = Depends(get_db)):
    """名刺削除"""
    card = db.query(BusinessCard).filter(BusinessCard.id == card_id).first()
    if not card:
        raise HTTPException(status_code=404, detail="Card not found")

    db.delete(card)
    db.commit()
    return {"deleted": card_id}


@app.get("/api/business-cards/stats/by-tag")
def get_cards_stats_by_tag(db: Session = Depends(get_db)):
    """タグ別統計"""
    stats = db.query(
        BusinessCard.tag,
        func.count(BusinessCard.id).label("count")
    ).group_by(BusinessCard.tag).all()

    return {s.tag or "other": s.count for s in stats}


# ========== Quote Document API (見積書) ==========

@app.get("/api/quotes")
def get_all_quotes(db: Session = Depends(get_db)):
    """全見積書一覧を取得"""
    quotes = db.query(QuoteDocument).order_by(QuoteDocument.created_at.desc()).all()
    result = []
    for q in quotes:
        items = db.query(QuoteItem).filter(QuoteItem.quote_id == q.id).order_by(QuoteItem.seq).all()
        result.append({
            "id": q.id,
            "quote_no": q.quote_no,
            "title": q.title,
            "client_name": q.client_name,
            "issue_date": q.issue_date.isoformat() if q.issue_date else None,
            "valid_until": q.valid_until.isoformat() if q.valid_until else None,
            "subtotal": q.subtotal,
            "tax_amount": q.tax_amount,
            "total": q.total,
            "notes": q.notes,
            "status": q.status,
            "project_id": q.project_id,
            "created_at": q.created_at.isoformat() if q.created_at else None,
            "items": [
                {
                    "id": item.id,
                    "seq": item.seq,
                    "name": item.name,
                    "specification": item.specification,
                    "quantity": item.quantity,
                    "unit": item.unit,
                    "unit_price": item.unit_price,
                    "amount": item.amount
                } for item in items
            ]
        })
    return result


@app.get("/api/quotes/{quote_id}")
def get_quote(quote_id: int, db: Session = Depends(get_db)):
    """見積書詳細を取得"""
    quote = db.query(QuoteDocument).filter(QuoteDocument.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")

    items = db.query(QuoteItem).filter(QuoteItem.quote_id == quote_id).order_by(QuoteItem.seq).all()
    return {
        "id": quote.id,
        "quote_no": quote.quote_no,
        "title": quote.title,
        "client_name": quote.client_name,
        "issue_date": quote.issue_date.isoformat() if quote.issue_date else None,
        "valid_until": quote.valid_until.isoformat() if quote.valid_until else None,
        "subtotal": quote.subtotal,
        "tax_amount": quote.tax_amount,
        "total": quote.total,
        "notes": quote.notes,
        "status": quote.status,
        "project_id": quote.project_id,
        "items": [
            {
                "id": item.id,
                "seq": item.seq,
                "name": item.name,
                "specification": item.specification,
                "quantity": item.quantity,
                "unit": item.unit,
                "unit_price": item.unit_price,
                "amount": item.amount
            } for item in items
        ]
    }


@app.post("/api/quotes")
def create_quote(data: dict, db: Session = Depends(get_db)):
    """見積書を作成"""
    items_data = data.pop("items", [])

    # 見積番号を自動生成
    if not data.get("quote_no"):
        import time
        data["quote_no"] = f"Q-{int(time.time())}"

    # 金額計算
    subtotal = sum(item.get("amount", 0) for item in items_data)
    tax_amount = int(subtotal * 0.1)
    total = subtotal + tax_amount

    quote = QuoteDocument(
        quote_no=data.get("quote_no"),
        title=data.get("title", ""),
        client_name=data.get("client_name", ""),
        issue_date=data.get("issue_date"),
        valid_until=data.get("valid_until"),
        subtotal=subtotal,
        tax_amount=tax_amount,
        total=total,
        notes=data.get("notes", ""),
        status="draft"
    )
    db.add(quote)
    db.flush()

    # 明細を登録
    for seq, item in enumerate(items_data):
        quote_item = QuoteItem(
            quote_id=quote.id,
            seq=seq,
            name=item.get("name", ""),
            specification=item.get("specification", ""),
            quantity=item.get("quantity", 1),
            unit=item.get("unit", "式"),
            unit_price=item.get("unit_price", 0),
            amount=item.get("amount", 0)
        )
        db.add(quote_item)

    db.commit()
    db.refresh(quote)
    return {"id": quote.id, "quote_no": quote.quote_no, "total": total}


@app.put("/api/quotes/{quote_id}")
def update_quote(quote_id: int, data: dict, db: Session = Depends(get_db)):
    """見積書を更新"""
    quote = db.query(QuoteDocument).filter(QuoteDocument.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")

    items_data = data.pop("items", [])

    # 既存明細を削除
    db.query(QuoteItem).filter(QuoteItem.quote_id == quote_id).delete()

    # 金額計算
    subtotal = sum(item.get("amount", 0) for item in items_data)
    tax_amount = int(subtotal * 0.1)
    total = subtotal + tax_amount

    # 見積書更新
    quote.title = data.get("title", quote.title)
    quote.client_name = data.get("client_name", quote.client_name)
    quote.issue_date = data.get("issue_date", quote.issue_date)
    quote.valid_until = data.get("valid_until", quote.valid_until)
    quote.notes = data.get("notes", quote.notes)
    quote.subtotal = subtotal
    quote.tax_amount = tax_amount
    quote.total = total

    # 明細を再登録
    for seq, item in enumerate(items_data):
        quote_item = QuoteItem(
            quote_id=quote.id,
            seq=seq,
            name=item.get("name", ""),
            specification=item.get("specification", ""),
            quantity=item.get("quantity", 1),
            unit=item.get("unit", "式"),
            unit_price=item.get("unit_price", 0),
            amount=item.get("amount", 0)
        )
        db.add(quote_item)

    db.commit()
    return {"id": quote.id, "total": total}


@app.delete("/api/quotes/{quote_id}")
def delete_quote(quote_id: int, db: Session = Depends(get_db)):
    """見積書を削除"""
    quote = db.query(QuoteDocument).filter(QuoteDocument.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")

    # 明細も削除
    db.query(QuoteItem).filter(QuoteItem.quote_id == quote_id).delete()
    db.delete(quote)
    db.commit()
    return {"deleted": quote_id}


@app.post("/api/quotes/{quote_id}/convert-to-order")
def convert_quote_to_order(quote_id: int, db: Session = Depends(get_db)):
    """見積書を受注に変換（工事・工種を自動作成）"""
    quote = db.query(QuoteDocument).filter(QuoteDocument.id == quote_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Quote not found")

    if quote.project_id:
        raise HTTPException(status_code=400, detail="Already converted to order")

    items = db.query(QuoteItem).filter(QuoteItem.quote_id == quote_id).order_by(QuoteItem.seq).all()

    # 工事コード生成
    import time
    from datetime import date
    project_code = f"P{date.today().strftime('%Y%m')}-{int(time.time()) % 10000:04d}"

    # 工事を作成
    project = Project(
        code=project_code,
        name=quote.title,
        client=quote.client_name,
        status="施工中",
        order_type="一次請",
        probability="確定",
        order_amount=quote.total,
        budget_amount=quote.subtotal,  # 税抜きを予算に
        tax_rate=0.1
    )
    db.add(project)
    db.flush()

    # 見積明細 → 工種として登録
    for seq, item in enumerate(items):
        work_type = ProjectWorkType(
            project_id=project.id,
            seq=seq + 1,
            name=item.name,
            dimension=item.specification or "",
            design_qty=item.quantity,
            unit=item.unit,
            budget_unit_price=item.unit_price,
            budget_amount=item.amount,
            rate=1.0
        )
        db.add(work_type)

    # 見積書のステータスを更新
    quote.status = "ordered"
    quote.project_id = project.id

    db.commit()

    return {
        "success": True,
        "project_id": project.id,
        "project_code": project_code,
        "project_name": project.name,
        "order_amount": quote.total,
        "work_types_count": len(items)
    }


# ============================================
# 見積書Excel出力機能
# ============================================

# スタイル定義
THIN = Side(style='thin', color='000000')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BLACK_FILL = PatternFill('solid', fgColor='000000')
YELLOW_FILL = PatternFill('solid', fgColor='FFFF99')
WHITE_FONT = Font(color='FFFFFF', bold=True)
TITLE_FONT = Font(size=16, bold=True)
SMALL_FONT = Font(size=9)
RED_FONT = Font(size=10, color='FF0000')


def create_cover_sheet(ws, data):
    """御見積書（表紙）シート作成"""
    col_widths = {'A': 10, 'B': 12, 'C': 18, 'D': 10, 'E': 6, 'F': 6, 'G': 12, 'H': 10, 'I': 12}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.merge_cells('A1:I1')
    ws['A1'] = "御 見 積 書"
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = BLACK_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"{data['estimate_info']['to_company']}　御中"
    ws['A2'].font = Font(size=11, bold=True)
    
    company = data['company_info']
    ws['G2'] = company['name']
    ws['G2'].font = Font(size=11, bold=True)
    ws['G3'] = company['postal']
    ws['G4'] = company['address']
    ws['G5'] = f"{company['tel']}  {company['fax']}"
    
    info = data['estimate_info']
    row = 7
    for label, value in [("【工 事 名】", info.get('project_name', '')),
                         ("【工事場所】", info.get('project_location', '')),
                         ("【工　　期】", info.get('period', '')),
                         ("【支払条件】", info.get('payment_terms', '')),
                         ("【受渡条件】", info.get('delivery_terms', '')),
                         ("【担 当 者】", info.get('contact', ''))]:
        ws[f'A{row}'] = label
        ws.merge_cells(f'B{row}:E{row}')
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "毎度、格別の御引立を賜り有難うございます。"
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "御依頼を戴きました本件に付き、誠心誠意検討を加え御見積申し上げましたので"
    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'] = "是非御下命賜ります様、お願い申し上げます。"
    
    row += 2
    subtotal = info.get('subtotal', 0)
    tax_rate = info.get('tax_rate', 0.10)
    tax = int(subtotal * tax_rate)
    total = subtotal + tax
    
    ws[f'C{row}'] = "小 計 金 額"
    ws[f'F{row}'] = f"¥{subtotal:,}"
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1
    ws[f'C{row}'] = f"消費税({int(tax_rate*100)}%)"
    ws[f'F{row}'] = f"¥{tax:,}"
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    row += 1
    ws.merge_cells(f'C{row}:D{row}')
    ws[f'C{row}'] = "合 計 金 額"
    ws[f'C{row}'].font = Font(size=12, bold=True)
    ws.merge_cells(f'F{row}:G{row}')
    ws[f'F{row}'] = f"¥{total:,}"
    ws[f'F{row}'].font = Font(size=14, bold=True)
    ws[f'F{row}'].alignment = Alignment(horizontal='right')
    
    row += 2
    headers = ["No.", "名　称", "仕様・規格・寸法", "設計", "数量", "単位", "金額", "単価", "備考"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = BLACK_FILL
        cell.font = Font(size=9, color='FFFFFF', bold=True)
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    for idx, wt in enumerate(data['work_types'], 1):
        ws.cell(row=row, column=1, value=idx).border = BORDER_ALL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=wt['name']).border = BORDER_ALL
        ws.cell(row=row, column=3, value=wt.get('spec', '') or "内訳書別添え").border = BORDER_ALL
        ws.cell(row=row, column=4, value="").border = BORDER_ALL
        ws.cell(row=row, column=5, value=wt.get('quantity', 1)).border = BORDER_ALL
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6, value=wt.get('unit', '式')).border = BORDER_ALL
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=7, value=wt.get('amount', 0)).border = BORDER_ALL
        ws.cell(row=row, column=7).number_format = '#,##0'
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=8, value="").border = BORDER_ALL
        ws.cell(row=row, column=9, value="").border = BORDER_ALL
        row += 1
    
    for _ in range(2):
        for i in range(1, 10):
            ws.cell(row=row, column=i, value="").border = BORDER_ALL
        row += 1
    
    ws.cell(row=row, column=1, value="").border = BORDER_ALL
    ws.merge_cells(f'B{row}:F{row}')
    ws.cell(row=row, column=2, value="合　計（税抜）").border = BORDER_ALL
    ws.cell(row=row, column=2).font = Font(bold=True)
    for i in range(3, 7):
        ws.cell(row=row, column=i).border = BORDER_ALL
    ws.cell(row=row, column=7, value=subtotal).border = BORDER_ALL
    ws.cell(row=row, column=7).number_format = '#,##0'
    ws.cell(row=row, column=7).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=7).font = Font(bold=True)
    ws.cell(row=row, column=8, value="").border = BORDER_ALL
    ws.cell(row=row, column=9, value="").border = BORDER_ALL


def create_breakdown_sheet(ws, work_type, sheet_num, company_info):
    """内訳明細書シート作成"""
    col_widths = {'A': 18, 'B': 22, 'C': 8, 'D': 6, 'E': 10, 'F': 12, 'G': 18}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    ws.merge_cells('A1:G1')
    ws['A1'] = "内 訳 明 細 書"
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = BLACK_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    row = 3
    headers = ["名　称", "規　格", "数量", "単位", "単価", "金額", "備　考"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.fill = YELLOW_FILL
        cell.font = Font(size=9, bold=True)
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = work_type['name']
    ws[f'A{row}'].fill = YELLOW_FILL
    ws[f'A{row}'].font = Font(size=10, bold=True)
    ws[f'A{row}'].border = BORDER_ALL
    for i in range(2, 8):
        ws.cell(row=row, column=i).border = BORDER_ALL
    
    row += 1
    category = work_type.get('category', '')
    if category:
        ws[f'A{row}'] = category
        ws[f'A{row}'].font = Font(size=9)
    
    row += 1
    items = work_type.get('items', [])
    direct_cost = 0
    
    for item in items:
        ws.cell(row=row, column=1, value=item.get('name', '')).border = BORDER_ALL
        ws.cell(row=row, column=2, value=item.get('spec', '')).border = BORDER_ALL
        qty = item.get('quantity', 0)
        ws.cell(row=row, column=3, value=qty if qty else '').border = BORDER_ALL
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4, value=item.get('unit', '')).border = BORDER_ALL
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        unit_price = item.get('unit_price', 0)
        ws.cell(row=row, column=5, value=unit_price if unit_price else '').border = BORDER_ALL
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        amount = item.get('amount', 0)
        ws.cell(row=row, column=6, value=amount if amount else '').border = BORDER_ALL
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7, value=item.get('note', '')).border = BORDER_ALL
        direct_cost += amount
        row += 1
    
    for _ in range(2):
        for i in range(1, 8):
            ws.cell(row=row, column=i, value="").border = BORDER_ALL
        row += 1
    
    summary = work_type.get('summary', {})
    
    # 直接工事費
    ws.cell(row=row, column=1, value="直接工事費").border = BORDER_ALL
    for i in range(2, 5):
        ws.cell(row=row, column=i, value="").border = BORDER_ALL
    ws.cell(row=row, column=3, value=1.0).border = BORDER_ALL
    ws.cell(row=row, column=4, value="式").border = BORDER_ALL
    ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    direct = summary.get('direct_cost', direct_cost)
    ws.cell(row=row, column=5, value="").border = BORDER_ALL
    ws.cell(row=row, column=6, value=direct).border = BORDER_ALL
    ws.cell(row=row, column=6).number_format = '#,##0'
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=7, value="").border = BORDER_ALL
    row += 1
    
    # 機械回送費
    transport = summary.get('transport_cost', 0)
    if transport:
        ws.cell(row=row, column=1, value="機械回送費").border = BORDER_ALL
        ws.cell(row=row, column=2, value=summary.get('transport_note', '')).border = BORDER_ALL
        ws.cell(row=row, column=3, value=1.0).border = BORDER_ALL
        ws.cell(row=row, column=4, value="往復").border = BORDER_ALL
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=transport).border = BORDER_ALL
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6, value=transport).border = BORDER_ALL
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7, value="").border = BORDER_ALL
        row += 1
    
    # 諸経費
    overhead = summary.get('overhead', 0)
    if overhead:
        ws.cell(row=row, column=1, value="諸　経　費").border = BORDER_ALL
        ws.cell(row=row, column=2, value=summary.get('overhead_note', '')).border = BORDER_ALL
        ws.cell(row=row, column=3, value=1.0).border = BORDER_ALL
        ws.cell(row=row, column=4, value="式").border = BORDER_ALL
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=overhead).border = BORDER_ALL
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6, value=overhead).border = BORDER_ALL
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7, value="").border = BORDER_ALL
        row += 1
    
    # 値引き
    discount = summary.get('discount', 0)
    if discount:
        ws.cell(row=row, column=1, value="値 引 き").border = BORDER_ALL
        ws.cell(row=row, column=2, value="").border = BORDER_ALL
        ws.cell(row=row, column=3, value=1.0).border = BORDER_ALL
        ws.cell(row=row, column=4, value="式").border = BORDER_ALL
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=abs(discount)).border = BORDER_ALL
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6, value=discount).border = BORDER_ALL
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6).font = RED_FONT
        ws.cell(row=row, column=7, value="").border = BORDER_ALL
        row += 1
    
    # 法定福利費
    welfare = summary.get('welfare_base', 0)
    welfare_rate = summary.get('welfare_rate', 0.15938)
    if welfare:
        ws.cell(row=row, column=1, value="法定福利費").border = BORDER_ALL
        ws.cell(row=row, column=2, value="事業主負担分").border = BORDER_ALL
        ws.cell(row=row, column=3, value=1.0).border = BORDER_ALL
        ws.cell(row=row, column=4, value="式").border = BORDER_ALL
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=welfare).border = BORDER_ALL
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=6, value=welfare).border = BORDER_ALL
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=7, value=f"社会保険料率\n〜{welfare_rate*100:.3f}%").border = BORDER_ALL
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)
        row += 1
    
    # 小計
    ws.cell(row=row, column=1, value="小　計").border = BORDER_ALL
    ws.cell(row=row, column=1).font = Font(size=10, bold=True)
    for i in range(2, 6):
        ws.cell(row=row, column=i, value="").border = BORDER_ALL
    subtotal = summary.get('subtotal', work_type.get('amount', 0))
    ws.cell(row=row, column=6, value=subtotal).border = BORDER_ALL
    ws.cell(row=row, column=6).number_format = '#,##0'
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=6).font = Font(size=10, bold=True)
    ws.cell(row=row, column=7, value="").border = BORDER_ALL


def create_condition_sheet(ws, conditions, company_info):
    """施工条件書シート作成"""
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 90
    
    ws.merge_cells('A1:B1')
    ws['A1'] = "施 工 条 件 書"
    ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
    ws['A1'].fill = BLACK_FILL
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    row = 3
    for i, cond in enumerate(conditions, 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=2, value=cond)
        ws.row_dimensions[row].height = 18
        row += 1
    
    row += 2
    ws.cell(row=row, column=2, value=company_info['name'])
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='right')


def create_estimate_excel_v2(data, output_path):
    """見積書Excelを作成"""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "御見積書"
    create_cover_sheet(ws1, data)
    
    for idx, wt in enumerate(data['work_types'], 1):
        ws = wb.create_sheet(f"内訳明細書{idx}")
        create_breakdown_sheet(ws, wt, idx, data['company_info'])
    
    ws3 = wb.create_sheet("施工条件書")
    create_condition_sheet(ws3, data.get('conditions', []), data['company_info'])
    
    wb.save(output_path)
    return output_path


# ============================================
# 見積書Excel出力API
# ============================================

@app.post("/api/projects/{project_id}/export-estimate")
async def export_estimate(project_id: int, db: Session = Depends(get_db)):
    """案件の見積書をExcelで出力"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    work_types = db.query(ProjectWorkType).filter(
        ProjectWorkType.project_id == project_id
    ).order_by(ProjectWorkType.seq).all()
    
    company_info = {
        "name": "株式会社 サンユウテック",
        "postal": "〒816-0912",
        "address": "福岡県大野城市御笠川6丁目2-5",
        "tel": "TEL092-555-9211",
        "fax": "FAX092-555-9217"
    }
    
    work_types_data = []
    subtotal = 0
    for wt in work_types:
        details = db.query(WorkTypeDetail).filter(
            WorkTypeDetail.work_type_id == wt.id
        ).order_by(WorkTypeDetail.seq).all()
        
        items = [{"name": d.name, "spec": d.spec or "", "quantity": d.budget_quantity or 0,
                  "unit": d.unit or "", "unit_price": d.budget_unit_price or 0,
                  "amount": d.budget_amount or 0} for d in details]
        
        direct_cost = sum(d.budget_amount or 0 for d in details)
        amount = wt.estimate_amount or wt.budget_amount or direct_cost
        subtotal += amount
        
        work_types_data.append({
            "name": wt.name,
            "spec": wt.spec or "内訳書別添え",
            "quantity": wt.quantity or 1,
            "unit": wt.unit or "式",
            "amount": amount,
            "category": wt.note or "",
            "items": items,
            "summary": {"direct_cost": direct_cost, "subtotal": amount}
        })
    
    data = {
        "company_info": company_info,
        "estimate_info": {
            "to_company": project.client or "",
            "project_name": project.name,
            "project_location": project.address or "",
            "period": getattr(project, 'period', '') or "",
            "payment_terms": "出来高請負払 現金100%",
            "delivery_terms": "別途、工事経理確認書及び現場条件書による",
            "contact": "上原 拓",
            "subtotal": subtotal,
            "tax_rate": 0.10,
        },
        "work_types": work_types_data,
        "conditions": []
    }
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        create_estimate_excel_v2(data, tmp.name)
        tmp_path = tmp.name
    
    filename = f"見積書_{project.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return FileResponse(path=tmp_path, filename=filename,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================
# 見積書Excel取込API（強化版）
# ============================================

@app.post("/api/projects/import-estimate")
async def import_estimate(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """見積書Excelを読み込んで新規案件として登録（内訳明細書も含む）"""
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        content = await file.read()
        tmp.write(content)
        tmp_path = tmp.name

    try:
        wb = load_workbook(tmp_path, data_only=True)

        # ============ 1. 御見積書シートから基本情報取得 ============
        project_name = ""
        client = ""
        location = ""
        period = ""
        work_types_data = []

        if "御見積書" in wb.sheetnames:
            ws = wb["御見積書"]

            # 基本情報を探す
            for row in ws.iter_rows(min_row=1, max_row=30):
                for cell in row:
                    val = str(cell.value) if cell.value else ""

                    # 宛先（御中）
                    if "御中" in val:
                        client = val.replace("御中", "").replace("　", " ").strip()

                    # 工事名
                    if "工事名" in val.replace(" ", "").replace("　", ""):
                        # 次のセルを確認
                        for c in range(cell.column + 1, cell.column + 5):
                            next_cell = ws.cell(row=cell.row, column=c)
                            if next_cell.value:
                                project_name = str(next_cell.value).strip()
                                break

                    # 工事場所
                    if "工事場所" in val.replace(" ", ""):
                        for c in range(cell.column + 1, cell.column + 5):
                            next_cell = ws.cell(row=cell.row, column=c)
                            if next_cell.value:
                                location = str(next_cell.value).strip()
                                break

                    # 工期
                    if "工期" in val.replace(" ", "").replace("　", ""):
                        for c in range(cell.column + 1, cell.column + 5):
                            next_cell = ws.cell(row=cell.row, column=c)
                            if next_cell.value:
                                period = str(next_cell.value).strip()
                                break

            # 工種テーブルを探す
            in_table = False
            for row in ws.iter_rows(min_row=10, max_row=100):
                first_val = row[0].value

                # ヘッダー行を検出
                if first_val and str(first_val).strip().lower() in ["no.", "no", "番号"]:
                    in_table = True
                    continue

                # データ行
                if in_table and first_val:
                    try:
                        no = int(first_val)
                        name = str(row[1].value).strip() if row[1].value else ""

                        if name and "合計" not in name:
                            # 列インデックスを柔軟に
                            spec = ""
                            quantity = 1
                            unit = "式"
                            amount = 0

                            for i, cell in enumerate(row[2:], start=2):
                                val = cell.value
                                if val is None:
                                    continue

                                # 数値を金額として判定
                                if isinstance(val, (int, float)) and val > 100:
                                    amount = int(val)
                                elif isinstance(val, (int, float)) and val <= 100:
                                    quantity = float(val)
                                elif isinstance(val, str):
                                    if val in ["式", "m", "m2", "m3", "t", "台", "人", "日", "往復"]:
                                        unit = val
                                    elif len(val) > 2:
                                        spec = val

                            work_types_data.append({
                                "seq": no,
                                "name": name,
                                "spec": spec,
                                "quantity": quantity,
                                "unit": unit,
                                "amount": amount
                            })
                    except (ValueError, TypeError):
                        # 合計行などはスキップ
                        if first_val and "合計" in str(first_val):
                            in_table = False

        # ============ 2. 内訳明細書シートから明細取得 ============
        all_details = {}  # 工種名 -> 明細リスト

        for sheet_name in wb.sheetnames:
            if "内訳" in sheet_name or "明細" in sheet_name:
                ws = wb[sheet_name]

                current_work_type = ""
                details = []

                for row in ws.iter_rows(min_row=1, max_row=300):
                    row_values = [cell.value for cell in row]
                    first_val = str(row[0].value).strip() if row[0].value else ""

                    # 工種名を検出（黄色背景 or 結合セル）
                    if row[0].fill and hasattr(row[0].fill, 'fgColor'):
                        fill_color = row[0].fill.fgColor
                        if fill_color and fill_color.rgb and 'FFFF' in str(fill_color.rgb):
                            # 黄色っぽい背景 = 工種名
                            if first_val and first_val not in ["名称", "名　称"]:
                                current_work_type = first_val
                                continue

                    # ヘッダー行をスキップ
                    if first_val in ["名称", "名　称", "品名"]:
                        continue

                    # 集計行をスキップ
                    skip_keywords = ["直接工事費", "機械回送費", "諸経費", "諸　経　費",
                                    "値引き", "値 引 き", "法定福利費", "小計", "小　計"]
                    if any(kw in first_val for kw in skip_keywords):
                        continue

                    # 空行スキップ
                    if not first_val:
                        continue

                    # 明細行として処理
                    try:
                        name = first_val
                        spec = str(row[1].value).strip() if len(row) > 1 and row[1].value else ""
                        quantity = 0
                        unit = ""
                        unit_price = 0
                        amount = 0

                        # 数量
                        if len(row) > 2 and row[2].value:
                            try:
                                quantity = float(row[2].value)
                            except:
                                pass

                        # 単位
                        if len(row) > 3 and row[3].value:
                            unit = str(row[3].value).strip()

                        # 単価
                        if len(row) > 4 and row[4].value:
                            try:
                                unit_price = int(float(row[4].value))
                            except:
                                pass

                        # 金額
                        if len(row) > 5 and row[5].value:
                            try:
                                amount = int(float(row[5].value))
                            except:
                                pass

                        # 有効なデータのみ追加
                        if name and (quantity > 0 or amount > 0):
                            detail = {
                                "name": name,
                                "spec": spec,
                                "quantity": quantity,
                                "unit": unit,
                                "unit_price": unit_price,
                                "amount": amount
                            }
                            details.append(detail)
                    except Exception as e:
                        print(f"明細行パースエラー: {e}")
                        continue

                # 工種名が特定できなかった場合はシート名から推測
                if not current_work_type:
                    current_work_type = sheet_name.replace("内訳明細書", "").strip() or "明細"

                if details:
                    all_details[current_work_type] = details

        # ============ 3. データベースに保存 ============

        # プロジェクト名がなければ生成
        if not project_name:
            project_name = f"取込案件_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        # 案件作成
        project = Project(
            name=project_name,
            client=client,
            address=location,
            status="見積中"
        )
        db.add(project)
        db.commit()
        db.refresh(project)

        # 工種が見つからなかった場合、内訳明細書のキーから作成
        if not work_types_data and all_details:
            for idx, (wt_name, details) in enumerate(all_details.items(), 1):
                total_amount = sum(d.get("amount", 0) for d in details)
                work_types_data.append({
                    "seq": idx,
                    "name": wt_name,
                    "spec": "内訳書別添え",
                    "quantity": 1,
                    "unit": "式",
                    "amount": total_amount
                })

        # 工種作成
        created_work_types = []
        for wt in work_types_data:
            work_type = ProjectWorkType(
                project_id=project.id,
                seq=wt.get("seq", 1),
                name=wt["name"],
                spec=wt.get("spec", ""),
                quantity=wt.get("quantity", 1),
                unit=wt.get("unit", "式"),
                budget_amount=wt.get("amount", 0),
                estimate_amount=wt.get("amount", 0),
                rate=1.0
            )
            db.add(work_type)
            db.commit()
            db.refresh(work_type)
            created_work_types.append(work_type)

        # 明細作成
        total_details = 0
        for wt in created_work_types:
            # 工種名で明細を検索
            details = all_details.get(wt.name, [])

            # 見つからなければ部分一致で検索
            if not details:
                for key in all_details.keys():
                    if wt.name in key or key in wt.name:
                        details = all_details[key]
                        break

            # それでもなければ最初の明細を使用（工種が1つの場合）
            if not details and len(created_work_types) == 1 and all_details:
                details = list(all_details.values())[0]

            for idx, d in enumerate(details, 1):
                detail = WorkTypeDetail(
                    work_type_id=wt.id,
                    seq=idx,
                    name=d.get("name", ""),
                    spec=d.get("spec", ""),
                    budget_quantity=d.get("quantity", 0),
                    unit=d.get("unit", ""),
                    budget_unit_price=d.get("unit_price", 0),
                    budget_amount=d.get("amount", 0),
                    cost_category="経費"  # デフォルト
                )
                db.add(detail)
                total_details += 1

            # 工種の予算金額を再計算
            if details:
                wt.budget_amount = sum(d.get("amount", 0) for d in details)
                wt.estimate_amount = wt.budget_amount

        db.commit()

        return {
            "success": True,
            "project_id": project.id,
            "project_name": project_name,
            "client": client,
            "location": location,
            "work_types_count": len(created_work_types),
            "details_count": total_details,
            "message": f"案件「{project_name}」を作成しました。工種{len(created_work_types)}件、明細{total_details}件を取り込みました。"
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"取込エラー: {str(e)}")

    finally:
        os.unlink(tmp_path)


# ============================================
# メンバー管理API
# ============================================

@app.get("/api/members")
def get_members(db: Session = Depends(get_db)):
    """メンバー一覧を取得"""
    members = db.query(Member).filter(Member.is_active == True).order_by(Member.id).all()
    return [{"id": m.id, "name": m.name, "department": m.department,
             "position": m.position, "line_works_id": m.line_works_id} for m in members]

@app.post("/api/members")
def create_member(data: dict, db: Session = Depends(get_db)):
    """メンバーを登録"""
    member = Member(
        name=data.get("name"),
        name_kana=data.get("name_kana"),
        email=data.get("email"),
        phone=data.get("phone"),
        department=data.get("department"),
        position=data.get("position"),
        line_works_id=data.get("line_works_id"),
        is_active=True
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return {"success": True, "id": member.id, "name": member.name}

@app.post("/api/members/init")
def init_members(db: Session = Depends(get_db)):
    """初期メンバーデータを登録"""
    initial_members = [
        {"name": "上原 拓", "department": "工事部", "position": "代表"},
        {"name": "田中 太郎", "department": "工事部", "position": "現場監督"},
        {"name": "山田 次郎", "department": "工事部", "position": "作業員"},
        {"name": "佐藤 花子", "department": "事務", "position": "経理"},
        {"name": "鈴木 一郎", "department": "工事部", "position": "作業員"},
    ]

    count = 0
    for m in initial_members:
        existing = db.query(Member).filter(Member.name == m["name"]).first()
        if not existing:
            member = Member(**m, is_active=True)
            db.add(member)
            count += 1

    db.commit()
    return {"success": True, "added": count, "message": f"{count}名のメンバーを登録しました"}


# ============================================
# ホテル予約依頼API
# ============================================

@app.post("/api/hotel-request")
async def create_hotel_request(data: dict, db: Session = Depends(get_db)):
    """ホテル予約依頼を作成し、LINE WORKSに送信"""
    import json

    # 依頼データを保存
    members_list = data.get("members", [])
    hotel_request = HotelRequest(
        project_id=data.get("project_id"),
        project_name=data.get("project_name"),
        location=data.get("location"),
        checkin_date=datetime.strptime(data.get("checkin"), "%Y-%m-%d").date() if data.get("checkin") else None,
        checkout_date=datetime.strptime(data.get("checkout"), "%Y-%m-%d").date() if data.get("checkout") else None,
        nights=data.get("nights", 1),
        members=json.dumps(members_list, ensure_ascii=False),
        member_count=len(members_list),
        status="pending",
        notes=data.get("notes"),
        requested_by=data.get("requested_by", "システム")
    )
    db.add(hotel_request)
    db.commit()
    db.refresh(hotel_request)

    # LINE WORKSメッセージを作成
    checkin_date = data.get("checkin", "")
    checkout_date = data.get("checkout", "")
    nights = data.get("nights", 1)

    # 日付フォーマット
    def format_date(date_str):
        if not date_str:
            return ""
        try:
            d = datetime.strptime(date_str, "%Y-%m-%d")
            weekdays = ["月", "火", "水", "木", "金", "土", "日"]
            return f"{d.month}/{d.day}({weekdays[d.weekday()]})"
        except:
            return date_str

    checkin_fmt = format_date(checkin_date)
    checkout_fmt = format_date(checkout_date)

    # メンバーリスト
    members_text = "\n".join([f"  ・{m}" for m in members_list]) if members_list else "  ・（未指定）"

    # 検索リンク
    search_links = data.get("search_links", {})
    links_text = ""
    if search_links:
        links_text = "\n".join([f"・{name}: {url}" for name, url in search_links.items()])

    message = f"""🏨 ホテル予約依頼

📍 現場：{data.get("project_name", "未指定")}
📍 場所：{data.get("location", "未指定")}
📅 日程：{checkin_fmt} → {checkout_fmt} {nights}泊
👥 メンバー：
{members_text}

🔗 最安値検索リンク
{links_text}

---
依頼ID: #{hotel_request.id}"""

    # LINE WORKS設定を取得して送信
    lw_settings = db.query(LineWorksSettings).first()
    lw_sent = False

    if lw_settings and lw_settings.is_active:
        try:
            # LINE WORKSへ送信（Bot API使用）
            # ここでは通知ログとして保存
            notification = LineWorksNotification(
                type="hotel_request",
                title="ホテル予約依頼",
                message=message[:500],
                target_type="bot",
                status="pending"
            )
            db.add(notification)

            # ログ記録
            log = LineWorksLog(
                action="send_message",
                target_type="bot",
                status="success",
                request_data=json.dumps({"message": message[:200]}, ensure_ascii=False),
                response_data=json.dumps({"hotel_request_id": hotel_request.id})
            )
            db.add(log)
            db.commit()
            lw_sent = True
        except Exception as e:
            print(f"LINE WORKS送信エラー: {e}")

    return {
        "success": True,
        "request_id": hotel_request.id,
        "message": message,
        "line_works_sent": lw_sent
    }

@app.get("/api/hotel-requests")
def get_hotel_requests(db: Session = Depends(get_db)):
    """ホテル予約依頼一覧を取得"""
    requests = db.query(HotelRequest).order_by(HotelRequest.created_at.desc()).limit(50).all()
    return [{
        "id": r.id,
        "project_name": r.project_name,
        "location": r.location,
        "checkin_date": str(r.checkin_date) if r.checkin_date else None,
        "checkout_date": str(r.checkout_date) if r.checkout_date else None,
        "nights": r.nights,
        "member_count": r.member_count,
        "status": r.status,
        "created_at": str(r.created_at)
    } for r in requests]
